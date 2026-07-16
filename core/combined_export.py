"""
Combines multiple increments (each already loaded via
ui.mock_data.MockDataStore.get_increment_for_display, at whichever
version the user picked per increment) into ONE .xlsx with the same
four sheet names a single-increment export uses -- All Data, Sum Data,
Report, Changes -- for portfolio-level reporting across a project
instead of one file per increment.

All Data / Sum Data: every selected increment's rows, concatenated in
the SAME order the increments appear in the on-screen increment list
(not the order their checkboxes happened to be clicked -- simpler to
implement and more predictable for the user, since it matches what
they're already looking at top to bottom), with a new "Increment"
column prepended identifying which increment each row came from. This
resurrects the real source template's own Inc# column, which this app
has never populated until now since it only ever handled one increment
at a time. Immediately after each increment's own block of rows, a
per-increment SUBTOTAL row -- computed by calling the exact same
single-increment totals functions (core.excel_reader.all_data_totals/
live_sum_data_totals) once per increment, not a new calculation, see
build_combined_all_data_rows/build_combined_sum_data_rows -- styled
distinctly (bold + a light tint) from both regular rows and the final
Grand Total. One combined GRAND total row at the very bottom, same
convention as core.excel_export's single-increment totals row,
covering every included increment together, completely UNCHANGED by
the addition of subtotals (see combine_all_data_totals -- summing each
increment's own already-computed totals is equivalent to, and simpler
than, recomputing from raw combined rows -- and real_rows(), which
per-increment subtotal marker rows are filtered out by before any
grand-total recomputation from a flat row list, e.g. Sum Data's live
totals).

Report: each selected increment's existing grouped/totaled Report,
completely unchanged (including its own per-increment "Grand Total"
row), as a separate, visibly labeled section in sequence -- followed by
ONE overall "Grand Total (All Increments)" row summing every section
together at the very end. Deliberately NOT labeled just "Grand Total"
like the per-section rows, so it's never ambiguous which total is the
combined one when scanning the sheet.

Changes: State Revision Log, Update History, AND Comments (see
core.excel_export._write_changes_sheet, the single-increment version of
this sheet) are each flattened across every selected increment with a
leading Increment column -- the SAME "concatenate, don't section"
treatment as All Data/Sum Data, deliberately NOT Report's per-increment
sectioning. That choice was made explicitly (not just for consistency):
none of the three Changes sub-tables has a per-increment OR overall
total worth visually separating for the way Report's sections exist
specifically to set off its per-section and Grand Total rows --
sectioning here would only cost readability/filterability (a flat table
sorts and filters in Excel; per-increment blocks don't) without buying
back anything sectioning is for. Each increment's own rows stay
contiguous, in on-screen order, exactly like All Data/Sum Data; Update
History and Comments are additionally newest-first WITHIN each
increment's block (matching core.excel_export's single-increment
ordering), not globally interleaved by date across increments -- a user
wanting a true cross-increment timeline can sort the Date column
themselves, which this flattened layout enables and per-increment
sectioning would not have. No combined totals row (there is nothing to
sum).

Column layout is shifted +1 versus core/excel_export.py throughout (the
new leading Increment column), so this module keeps its own column
constants and small helper functions rather than reusing
core.excel_export's directly -- those hardcode positions that would
misalign by one column here. Style constants (colors/fonts) and the
small Changes-sheet formatting helpers ARE reused from there, so this
never invents a second, subtly-different palette or phrasing.

======================================================================
SHARED ROW-BUILDING FUNCTIONS (build_combined_*, combine_all_data_totals)
======================================================================
The public build_combined_*/combine_all_data_totals functions below are
the SINGLE source of truth for which rows appear, in what order, and
with what derived values (live status, counts, section/grand-total
flags) -- pure Python, no openpyxl. The _write_combined_*_sheet
functions consume them and add nothing but cell styling/formatting.
ui.mock_data.build_combined_view() consumes the exact same functions to
build the on-screen combined preview (ui/pages/combined_data_view_page.py).
This is deliberate, not incidental: it is the mechanism that guarantees
the preview and the exported .xlsx can never disagree about content --
there is no second, independently-maintained copy of "which rows, what
order, what totals" for the UI to drift from.
"""

from __future__ import annotations

from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from core.excel_export import (
    _BOLD_FONT,
    _DESCRIPTION_ALIGNMENT,
    _HEADER_FONT,
    _STAGE_ALIGNMENT,
    _STAGE_FILLS,
    _STAGE_FONT,
    _STAGE_MARKS,
    _SUM_DATA_STAGE_FILLS,
    _SUM_DATA_STAGE_FONTS,
    _SUM_DATA_STAGE_LABELS,
    _TOTALS_FILL,
    _TOTALS_FONT,
    _UNSAFE_FILENAME_CHARS,
    _WHITESPACE,
    _apply_description_wrap,
    _format_changes_date,
    _history_entry_detail_text,
    _history_entry_summary,
    _wrapped_row_height,
)
from core.excel_reader import STAGE_COUNT, live_sum_data_totals

INCREMENT_COL = 1
INDEX_COL = 2
DESCRIPTION_COL = 3
AGENCY_COL = 4
STAGE_FIRST_COL = 5
STAGE_LAST_COL = STAGE_FIRST_COL + STAGE_COUNT - 1

# Per-increment subtotal row -- bold, like the Grand Total row (_TOTALS_FONT/
# _TOTALS_FILL, imported above), but a deliberately LIGHTER tint (#eef1f7 vs
# _TOTALS_FILL's #e4e8f0) so the row hierarchy -- regular row < subtotal <
# Grand Total -- reads clearly at a glance, in both the export and the
# CombinedDataViewPage UI (see ui/pages/combined_data_view_page.py's matching
# SUBTOTAL_ROW_BACKGROUND, the same color).
_SUBTOTAL_FILL = PatternFill(start_color="FFEEF1F7", end_color="FFEEF1F7", fill_type="solid")
_SUBTOTAL_FONT = Font(bold=True)

# --- Changes sheet column layout -- see _write_combined_changes_sheet.
# Same +1 shift (leading Increment column) versus
# core.excel_export._write_changes_sheet's own column constants; the two
# mini-tables share this one 7-column grid (columns 3-7 pair up the SAME
# way core.excel_export's single-increment version pairs its 6 columns --
# e.g. col 3 holds Synopsis in one table and Detail in the other, both
# long text -- just shifted +1 here).
CHANGES_REVISION_NUM_COL = 2
CHANGES_SYNOPSIS_COL = 3
CHANGES_AOR_SIGNATURE_COL = 4
CHANGES_SEOR_SIGNATURE_COL = 5
CHANGES_EFFECTIVE_DATE_COL = 6
CHANGES_HCAI_CONCURRENCE_COL = 7

CHANGES_UPDATE_NUM_COL = 2
CHANGES_DETAIL_COL = 3
CHANGES_SUMMARY_COL = 4
CHANGES_OLD_VERSION_COL = 5
CHANGES_NEW_VERSION_COL = 6
CHANGES_UPDATE_DATE_COL = 7

CHANGES_COMMENT_NUM_COL = 2
CHANGES_COMMENT_TEXT_COL = 3
CHANGES_COMMENT_DATE_COL = 4


def default_combined_filename(project_name: str, increment_count: int) -> str:
    """"{ProjectName}_Combined_{N}-increments.xlsx" -- same filesystem-
    unsafe-character sanitization as core.excel_export.default_filename.
    """

    def clean(text: str) -> str:
        stripped = _UNSAFE_FILENAME_CHARS.sub("", text)
        return _WHITESPACE.sub("_", stripped.strip())

    return f"{clean(project_name)}_Combined_{increment_count}-increments.xlsx"


def _stage_headers() -> list[str]:
    return [f"Stage {i}" for i in range(1, STAGE_COUNT + 1)]


def _write_header(ws: Worksheet, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT


def _write_totals_row(ws: Worksheet, values: list[Any]) -> None:
    ws.append(values)
    for cell in ws[ws.max_row]:
        cell.font = _TOTALS_FONT
        cell.fill = _TOTALS_FILL
        if cell.column > DESCRIPTION_COL:
            cell.alignment = _STAGE_ALIGNMENT


def _write_stage_cell(ws: Worksheet, excel_row: int, stage: int, status: str | None) -> None:
    cell = ws.cell(row=excel_row, column=STAGE_FIRST_COL + stage - 1, value=_STAGE_MARKS.get(status, ""))
    fill = _STAGE_FILLS.get(status)
    if fill is not None:
        cell.fill = fill
        cell.font = _STAGE_FONT
        cell.alignment = _STAGE_ALIGNMENT


def _write_sum_data_stage_cell(ws: Worksheet, excel_row: int, stage: int, status: str) -> None:
    cell = ws.cell(row=excel_row, column=STAGE_FIRST_COL + stage - 1, value=_SUM_DATA_STAGE_LABELS[status])
    cell.fill = _SUM_DATA_STAGE_FILLS[status]
    cell.font = _SUM_DATA_STAGE_FONTS[status]
    cell.alignment = _STAGE_ALIGNMENT


def combine_all_data_totals(increments: list[Any]) -> dict[str, Any]:
    """Sums each increment's OWN already-computed all_data_totals dict
    together, key by key. Correct and simpler than recomputing from raw
    combined rows from scratch: these are already per-column numeric
    sums (see core.excel_reader.all_data_totals), and addition is
    associative/commutative, so summing the per-increment totals gives
    exactly the same result.
    """
    combined: dict[str, Any] = {f"Stage {s}": 0 for s in range(1, STAGE_COUNT + 1)}
    combined["VCR"] = 0
    combined["SUM"] = 0
    for increment in increments:
        for key, value in increment.all_data_totals.items():
            combined[key] = combined.get(key, 0) + (value or 0)
    return combined


def real_rows(rows: list[dict]) -> list[dict]:
    """Filters out per-increment subtotal marker rows (is_subtotal=True),
    leaving only genuine data rows -- for any caller that needs to
    recompute an OVERALL total from a flat row list (e.g. Sum Data's
    live Grand Total, both here and in ui.mock_data.build_combined_view)
    and must not let a subtotal marker mixed into that same list throw
    the computation off. A subtotal marker actually contributes nothing
    even if left in (it carries no "required_stages"/"stage_status"), but
    filtering explicitly is clearer than relying on that incidental
    safety.
    """
    return [row for row in rows if not row.get("is_subtotal")]


def build_combined_all_data_rows(increments: list[Any]) -> list[dict]:
    """[{"increment": name, "is_subtotal": False, **row_data}, ...] for
    every real row, PLUS one subtotal marker row
    ({"increment": name, "is_subtotal": True, "subtotal_totals": {...}})
    immediately after each increment's own block -- exactly the rows/
    order _write_combined_all_data_sheet writes (see module docstring).

    "subtotal_totals" is that SAME increment's own already-computed
    core.excel_reader.all_data_totals() result (increment.all_data_totals,
    cached on the Increment at load time) -- not a new calculation, per
    the module docstring's reuse guarantee. The final combined Grand
    Total is unaffected by any of this: combine_all_data_totals() below
    sums increment.all_data_totals directly, never derived from this row
    list.
    """
    rows = []
    for increment in increments:
        for row_data in increment.all_data:
            rows.append({"increment": increment.name, **row_data, "is_subtotal": False})
        rows.append(
            {
                "increment": increment.name,
                "is_subtotal": True,
                "subtotal_totals": increment.all_data_totals,
            }
        )
    return rows


def build_combined_sum_data_rows(increments: list[Any]) -> list[dict]:
    """Same as build_combined_all_data_rows, but for Sum Data -- each
    REAL row dict additionally carries precomputed "live_status" (per
    required stage, defaulting any stage with no status.json entry to
    "Open" -- Sum Data's "current state" framing, same as
    core.excel_export's single-increment version), "open_count",
    "done_count", "total_count", "pct_complete", so nothing consuming
    this list ever independently re-derives (and risks disagreeing on)
    these live-status-dependent values.

    Each SUBTOTAL marker row's "subtotal_totals" is
    core.excel_reader.live_sum_data_totals() called on that ONE
    increment's own raw increment.sum_data list -- the exact same
    function the combined Grand Total (and the single-increment Sum
    Data tab) already calls, just scoped to one increment at a time
    instead of the flattened combined list, per the module docstring's
    reuse guarantee. Use real_rows() before calling
    live_sum_data_totals() on this function's own output for an OVERALL
    total -- these marker rows would otherwise sit alongside real ones
    in the same flat list.
    """
    rows = []
    for increment in increments:
        for row_data in increment.sum_data:
            required = row_data.get("required_stages", [])
            stage_status = row_data.get("stage_status", {})
            live_status = {stage: stage_status.get(stage, "Open") for stage in required}
            done_count = sum(1 for status in live_status.values() if status == "Done")
            open_count = sum(1 for status in live_status.values() if status == "Open")
            total_count = len(required)
            pct = (done_count / total_count) if required else None
            rows.append(
                {
                    "increment": increment.name,
                    **row_data,
                    "is_subtotal": False,
                    "live_status": live_status,
                    "open_count": open_count,
                    "done_count": done_count,
                    "total_count": total_count,
                    "pct_complete": pct,
                }
            )
        rows.append(
            {
                "increment": increment.name,
                "is_subtotal": True,
                "subtotal_totals": live_sum_data_totals(increment.sum_data),
            }
        )
    return rows


def build_combined_report_rows(increments: list[Any]) -> list[dict]:
    """The FULL Report row sequence -- per-increment section header rows,
    every item row, and the final combined Grand Total row -- everything
    _write_combined_report_sheet writes. Each row is explicitly tagged
    ("is_section_header"/"is_grand_total"/"is_combined_grand_total")
    rather than leaving a consumer to infer row type from field values
    (e.g. approval_agency == "Grand Total"), so a real item whose
    Approval Agency happens to literally be that string can never be
    misidentified as a totals row downstream.
    """
    rows = []
    combined_grand_total = 0.0
    for increment in increments:
        rows.append(
            {
                "is_section_header": True, "is_grand_total": False, "is_combined_grand_total": False,
                "approval_agency": increment.name, "index": "", "description": "", "total": None,
            }
        )
        for row_data in increment.report:
            is_grand_total = row_data.get("approval_agency") == "Grand Total"
            rows.append(
                {
                    "is_section_header": False, "is_grand_total": is_grand_total, "is_combined_grand_total": False,
                    "approval_agency": row_data.get("approval_agency") or "",
                    "index": row_data.get("index") or "",
                    "description": row_data.get("description") or "",
                    "total": row_data.get("total", 0),
                }
            )
            if is_grand_total:
                combined_grand_total += row_data.get("total", 0) or 0
    rows.append(
        {
            "is_section_header": False, "is_grand_total": False, "is_combined_grand_total": True,
            "approval_agency": "Grand Total (All Increments)", "index": "", "description": "",
            "total": combined_grand_total,
        }
    )
    return rows


def build_combined_changes_data(increments: list[Any]) -> dict[str, list[dict]]:
    """{"revision_log_rows": [...], "update_history_rows": [...]} --
    both flattened across increments with a leading "increment" tag, in
    the exact row set/order _write_combined_changes_sheet writes. Update
    History rows are newest-first WITHIN each increment's own block (see
    module docstring for why this isn't globally interleaved by date),
    and additionally carry a 1-based "update_number" per increment
    (matching the per-increment "Update #" column the export writes).
    """
    revision_log_rows = [
        {"increment": increment.name, **entry} for increment in increments for entry in increment.changes_log
    ]
    update_history_rows = [
        {"increment": increment.name, "update_number": i, **entry}
        for increment in increments
        for i, entry in enumerate(reversed(increment.change_history), start=1)
    ]
    return {"revision_log_rows": revision_log_rows, "update_history_rows": update_history_rows}


def build_combined_comments_rows(increments: list[Any]) -> list[dict]:
    """Comments flattened across every selected increment with a leading
    "increment" tag, newest-first WITHIN each increment's own block, each
    increment's own rows contiguous in on-screen order -- the exact row
    set/order _write_combined_changes_sheet writes for the Comments
    table (see module docstring for why this is flattened, not
    sectioned, same as State Revision Log/Update History). Carries a
    1-based "comment_number" per increment, matching the per-increment
    "#" column the export writes.
    """
    return [
        {"increment": increment.name, "comment_number": i, **entry}
        for increment in increments
        for i, entry in enumerate(reversed(increment.comments), start=1)
    ]


def _write_all_data_subtotal_row(ws: Worksheet, row_data: dict) -> None:
    """One increment's own subtotal row -- bold + a light tint
    (_SUBTOTAL_FILL), distinct from both regular rows (no fill) and the
    Grand Total row at the very bottom (_TOTALS_FILL, more saturated).
    "subtotal_totals" is that increment's OWN already-computed
    core.excel_reader.all_data_totals() result -- see
    build_combined_all_data_rows.
    """
    totals = row_data["subtotal_totals"]
    values = [f"{row_data['increment']} — Subtotal", "", "", ""]
    for stage in range(1, STAGE_COUNT + 1):
        values.append(totals.get(f"Stage {stage}", 0))
    values.append(totals.get("VCR", 0))
    values.append(totals.get("SUM", 0))
    ws.append(values)
    for cell in ws[ws.max_row]:
        cell.font = _SUBTOTAL_FONT
        cell.fill = _SUBTOTAL_FILL
        if cell.column > DESCRIPTION_COL:
            cell.alignment = _STAGE_ALIGNMENT


def _write_combined_all_data_sheet(ws: Worksheet, increments: list[Any]) -> None:
    headers = ["Increment", "Index", "Description", "Approval Agency"] + _stage_headers() + ["VCR", "SUM"]
    _write_header(ws, headers)

    for row_data in build_combined_all_data_rows(increments):
        if row_data["is_subtotal"]:
            _write_all_data_subtotal_row(ws, row_data)
            continue

        required = set(row_data.get("required_stages", []))
        stage_status = row_data.get("stage_status", {})
        values = [
            row_data["increment"],
            row_data.get("index"),
            row_data.get("description") or "",
            row_data.get("approval_agency") or "",
        ]
        for stage in range(1, STAGE_COUNT + 1):
            if stage in required:
                values.append(_STAGE_MARKS.get(stage_status.get(stage), ""))
            else:
                # Same "0, not blank" convention as core.excel_export.
                values.append(0)
        vcr = row_data.get("vcr")
        values.append(0 if vcr is None else vcr)
        values.append(row_data.get("sum"))
        ws.append(values)

        excel_row = ws.max_row
        _apply_description_wrap(ws, excel_row, column=DESCRIPTION_COL)
        for stage in required:
            _write_stage_cell(ws, excel_row, stage, stage_status.get(stage))
        for stage in range(1, STAGE_COUNT + 1):
            if stage not in required:
                ws.cell(row=excel_row, column=STAGE_FIRST_COL + stage - 1).alignment = _STAGE_ALIGNMENT

    # Grand Total -- UNCHANGED by the addition of per-increment subtotals
    # above: sums increment.all_data_totals directly, never derived from
    # the row list this loop just walked.
    totals = combine_all_data_totals(increments)
    totals_values = ["Totals", "", "", ""]
    for stage in range(1, STAGE_COUNT + 1):
        totals_values.append(totals.get(f"Stage {stage}", 0))
    totals_values.append(totals.get("VCR", 0))
    totals_values.append(totals.get("SUM", 0))
    _write_totals_row(ws, totals_values)

    vcr_col = STAGE_LAST_COL + 1
    sum_col = vcr_col + 1
    ws.column_dimensions[get_column_letter(INCREMENT_COL)].width = 40
    ws.column_dimensions[get_column_letter(INDEX_COL)].width = 12
    ws.column_dimensions[get_column_letter(DESCRIPTION_COL)].width = 45
    ws.column_dimensions[get_column_letter(AGENCY_COL)].width = 20
    for col in range(STAGE_FIRST_COL, STAGE_LAST_COL + 1):
        ws.column_dimensions[get_column_letter(col)].width = 9
    ws.column_dimensions[get_column_letter(vcr_col)].width = 9
    ws.column_dimensions[get_column_letter(sum_col)].width = 9
    ws.freeze_panes = ws.cell(row=2, column=STAGE_FIRST_COL).coordinate


def _write_sum_data_subtotal_row(ws: Worksheet, row_data: dict) -> None:
    """One increment's own subtotal row -- same bold + light-tint
    treatment as _write_all_data_subtotal_row. "subtotal_totals" is
    core.excel_reader.live_sum_data_totals() called on that ONE
    increment's own raw sum_data list -- see build_combined_sum_data_rows.
    """
    totals = row_data["subtotal_totals"]
    values = [f"{row_data['increment']} — Subtotal", "", "", ""]
    for stage in range(1, STAGE_COUNT + 1):
        values.append(totals["stage_open_counts"][stage])
    values.append(totals["vcr_open_count"])
    values.append(totals["open_total"])
    values.append(totals["done_total"])
    values.append(totals["grand_total"])
    values.append(totals["pct_complete"])
    ws.append(values)
    excel_row = ws.max_row
    for cell in ws[excel_row]:
        cell.font = _SUBTOTAL_FONT
        cell.fill = _SUBTOTAL_FILL
        if cell.column > DESCRIPTION_COL:
            cell.alignment = _STAGE_ALIGNMENT
    pct_col = STAGE_LAST_COL + 5
    if totals["pct_complete"] is not None:
        ws.cell(row=excel_row, column=pct_col).number_format = "0%"


def _write_combined_sum_data_sheet(ws: Worksheet, increments: list[Any]) -> None:
    headers = ["Increment", "Index", "Description", "Approval Agency"] + _stage_headers() + [
        "VCR", "Open", "Done", "Total", "% Complete"
    ]
    _write_header(ws, headers)

    rows = build_combined_sum_data_rows(increments)
    for row_data in rows:
        if row_data["is_subtotal"]:
            _write_sum_data_subtotal_row(ws, row_data)
            continue

        live_status = row_data["live_status"]
        pct = row_data["pct_complete"]

        values = [
            row_data["increment"],
            row_data.get("index"),
            row_data.get("description") or "",
            row_data.get("approval_agency") or "",
        ]
        values += ["" for _ in range(1, STAGE_COUNT + 1)]  # written properly below via _write_sum_data_stage_cell
        values.append(row_data.get("vcr"))
        values.append(row_data["open_count"])
        values.append(row_data["done_count"])
        values.append(row_data["total_count"])
        values.append(pct)
        ws.append(values)

        excel_row = ws.max_row
        _apply_description_wrap(ws, excel_row, column=DESCRIPTION_COL)
        for stage, status in live_status.items():
            _write_sum_data_stage_cell(ws, excel_row, stage, status)

        pct_col = STAGE_LAST_COL + 5
        if pct is not None:
            ws.cell(row=excel_row, column=pct_col).number_format = "0%"

    vcr_col = STAGE_LAST_COL + 1
    open_col = vcr_col + 1
    done_col = vcr_col + 2
    total_col = vcr_col + 3
    pct_col = vcr_col + 4

    # Live GRAND total across ALL selected increments' REAL rows together
    # -- recomputed fresh every export, same as the single-increment case
    # (see core.excel_reader.live_sum_data_totals), never cached, and
    # UNCHANGED by the addition of per-increment subtotals above:
    # real_rows() strips this function's own subtotal marker rows out of
    # `rows` first, so they can't be double-counted or otherwise throw
    # this off.
    totals = live_sum_data_totals(real_rows(rows))
    totals_values = ["Totals", "", "", ""]
    for stage in range(1, STAGE_COUNT + 1):
        totals_values.append(totals["stage_open_counts"][stage])
    totals_values.append(totals["vcr_open_count"])
    totals_values.append(totals["open_total"])
    totals_values.append(totals["done_total"])
    totals_values.append(totals["grand_total"])
    totals_values.append(totals["pct_complete"])
    _write_totals_row(ws, totals_values)
    if totals["pct_complete"] is not None:
        ws.cell(row=ws.max_row, column=pct_col).number_format = "0%"

    ws.column_dimensions[get_column_letter(INCREMENT_COL)].width = 40
    ws.column_dimensions[get_column_letter(INDEX_COL)].width = 12
    ws.column_dimensions[get_column_letter(DESCRIPTION_COL)].width = 45
    ws.column_dimensions[get_column_letter(AGENCY_COL)].width = 20
    for col in range(STAGE_FIRST_COL, STAGE_LAST_COL + 1):
        ws.column_dimensions[get_column_letter(col)].width = 9
    ws.column_dimensions[get_column_letter(vcr_col)].width = 9
    ws.column_dimensions[get_column_letter(open_col)].width = 9
    ws.column_dimensions[get_column_letter(done_col)].width = 9
    ws.column_dimensions[get_column_letter(total_col)].width = 9
    ws.column_dimensions[get_column_letter(pct_col)].width = 12
    ws.freeze_panes = ws.cell(row=2, column=STAGE_FIRST_COL).coordinate


def _write_combined_report_sheet(ws: Worksheet, increments: list[Any]) -> None:
    headers = ["Approval Agency", "Index", "Description", "Total"]
    _write_header(ws, headers)

    for row_data in build_combined_report_rows(increments):
        ws.append(
            [
                row_data["approval_agency"],
                row_data["index"],
                row_data["description"],
                row_data["total"] if row_data["total"] is not None else "",
            ]
        )
        excel_row = ws.max_row
        if row_data["is_section_header"] or row_data["is_combined_grand_total"]:
            # Visually distinct (bold + tint) -- reusing the same "this
            # is a summary, not a real item" styling the bottom totals
            # rows already use elsewhere in this app. Deliberately NOT
            # labeled just "Grand Total" like each section's own row, so
            # it's never ambiguous which total is the combined one when
            # scanning the sheet.
            for cell in ws[excel_row]:
                cell.font = _TOTALS_FONT
                cell.fill = _TOTALS_FILL
        else:
            _apply_description_wrap(ws, excel_row, column=3)
            if row_data["is_grand_total"]:
                for cell in ws[excel_row]:
                    cell.font = _BOLD_FONT

    ws.column_dimensions[get_column_letter(1)].width = 30
    ws.column_dimensions[get_column_letter(2)].width = 12
    ws.column_dimensions[get_column_letter(3)].width = 45
    ws.column_dimensions[get_column_letter(4)].width = 10


def _write_combined_changes_sheet(ws: Worksheet, increments: list[Any]) -> None:
    """State Revision Log, Update History, then Comments, each flattened
    across every selected increment with a leading Increment column --
    see the module docstring for why this deliberately does NOT use
    Report's per-increment sectioning.
    """
    section_row = 1
    ws.cell(row=section_row, column=1, value="State Revision Log").font = _BOLD_FONT

    header_row = section_row + 1
    revision_headers = [
        "Increment", "Rev #", "Synopsis of Change", "AOR Signature", "SEOR Signature", "Effective Date",
        "HCAI Concurrence",
    ]
    for col, label in enumerate(revision_headers, start=1):
        ws.cell(row=header_row, column=col, value=label).font = _HEADER_FONT

    changes_data = build_combined_changes_data(increments)
    revision_log_rows = changes_data["revision_log_rows"]
    update_history_rows = changes_data["update_history_rows"]

    row = header_row + 1
    for entry in revision_log_rows:
        synopsis = entry.get("synopsis") or ""
        ws.cell(row=row, column=INCREMENT_COL, value=entry["increment"])
        ws.cell(row=row, column=CHANGES_REVISION_NUM_COL, value=entry.get("revision_number"))
        ws.cell(row=row, column=CHANGES_SYNOPSIS_COL, value=synopsis).alignment = _DESCRIPTION_ALIGNMENT
        ws.cell(row=row, column=CHANGES_AOR_SIGNATURE_COL, value=_format_changes_date(entry.get("aor_signature_date")))
        ws.cell(row=row, column=CHANGES_SEOR_SIGNATURE_COL, value=_format_changes_date(entry.get("seor_signature_date")))
        ws.cell(row=row, column=CHANGES_EFFECTIVE_DATE_COL, value=_format_changes_date(entry.get("effective_date")))
        ws.cell(row=row, column=CHANGES_HCAI_CONCURRENCE_COL, value=entry.get("hcai_concurrence") or "")
        ws.row_dimensions[row].height = _wrapped_row_height(synopsis)
        row += 1
    if not revision_log_rows:
        ws.cell(row=row, column=1, value="No revision log entries found for any selected increment.")
        row += 1

    row += 1  # blank separator row
    history_section_row = row
    ws.cell(row=history_section_row, column=1, value="Update History").font = _BOLD_FONT

    history_header_row = history_section_row + 1
    history_headers = ["Increment", "Update #", "Detail", "Summary", "Old Version", "New Version", "Date"]
    for col, label in enumerate(history_headers, start=1):
        ws.cell(row=history_header_row, column=col, value=label).font = _HEADER_FONT

    row = history_header_row + 1
    for entry in update_history_rows:
        detail = _history_entry_detail_text(entry)
        ws.cell(row=row, column=INCREMENT_COL, value=entry["increment"])
        ws.cell(row=row, column=CHANGES_UPDATE_NUM_COL, value=entry["update_number"])
        ws.cell(row=row, column=CHANGES_DETAIL_COL, value=detail).alignment = _DESCRIPTION_ALIGNMENT
        ws.cell(row=row, column=CHANGES_SUMMARY_COL, value=_history_entry_summary(entry))
        ws.cell(row=row, column=CHANGES_OLD_VERSION_COL, value=entry.get("old_version"))
        ws.cell(row=row, column=CHANGES_NEW_VERSION_COL, value=entry.get("new_version"))
        ws.cell(row=row, column=CHANGES_UPDATE_DATE_COL, value=(entry.get("timestamp") or "").split("T", 1)[0])
        ws.row_dimensions[row].height = _wrapped_row_height(detail)
        row += 1
    if not update_history_rows:
        ws.cell(row=row, column=1, value="No updates confirmed yet for any selected increment.")
        row += 1

    row += 1  # blank separator row
    comments_section_row = row
    ws.cell(row=comments_section_row, column=1, value="Comments").font = _BOLD_FONT

    comments_header_row = comments_section_row + 1
    comments_headers = ["Increment", "#", "Comment", "Date"]
    for col, label in enumerate(comments_headers, start=1):
        ws.cell(row=comments_header_row, column=col, value=label).font = _HEADER_FONT

    comments_rows = build_combined_comments_rows(increments)
    row = comments_header_row + 1
    for entry in comments_rows:
        text = entry.get("text") or ""
        ws.cell(row=row, column=INCREMENT_COL, value=entry["increment"])
        ws.cell(row=row, column=CHANGES_COMMENT_NUM_COL, value=entry["comment_number"])
        ws.cell(row=row, column=CHANGES_COMMENT_TEXT_COL, value=text).alignment = _DESCRIPTION_ALIGNMENT
        ws.cell(row=row, column=CHANGES_COMMENT_DATE_COL, value=(entry.get("timestamp") or "").split("T", 1)[0])
        ws.row_dimensions[row].height = _wrapped_row_height(text)
        row += 1
    if not comments_rows:
        ws.cell(row=row, column=1, value="No comments yet for any selected increment.")

    ws.column_dimensions[get_column_letter(INCREMENT_COL)].width = 40
    ws.column_dimensions[get_column_letter(2)].width = 10
    ws.column_dimensions[get_column_letter(3)].width = 70
    ws.column_dimensions[get_column_letter(4)].width = 45
    ws.column_dimensions[get_column_letter(5)].width = 14
    ws.column_dimensions[get_column_letter(6)].width = 14
    ws.column_dimensions[get_column_letter(7)].width = 22


def export_combined_report(increments: list[Any], path: str) -> None:
    """Writes increments (a list of ui.mock_data.Increment, each already
    populated by MockDataStore.get_increment_for_display -- one entry
    per user-selected increment, at whichever version was picked for
    it, in the order they should appear in the output) to one .xlsx
    workbook with four combined sheets: All Data, Sum Data, Report,
    Changes.
    """
    wb = openpyxl.Workbook()
    ws_all = wb.active
    ws_all.title = "All Data"
    _write_combined_all_data_sheet(ws_all, increments)

    ws_sum = wb.create_sheet("Sum Data")
    _write_combined_sum_data_sheet(ws_sum, increments)

    ws_report = wb.create_sheet("Report")
    _write_combined_report_sheet(ws_report, increments)

    ws_changes = wb.create_sheet("Changes")
    _write_combined_changes_sheet(ws_changes, increments)

    wb.save(path)

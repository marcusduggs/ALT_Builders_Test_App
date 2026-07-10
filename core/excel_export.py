"""
Exports a cached ui.mock_data.Increment -- All Data, Sum Data, Report,
and Changes, all loaded together by
ui.mock_data.MockDataStore.get_increment_for_display() -- to a real
.xlsx workbook with four sheets, named and ordered exactly "All Data",
"Sum Data", "Report", "Changes", matching what's on screen in
ui.pages.data_view_page's four tabs (including any in-memory status
edits made this session that haven't been re-parsed from disk).

All Data's stage cell marks/colors mirror ui.pages.data_view_page's
_stage_status_icon() exactly ("X" on green = Done, "1" on red = Open)
using the same RGB values core.excel_reader._fill_color_status()
classifies as green/red, so re-uploading an exported file as a new
version is read back correctly by the app's own fill-color status
seeding (core.status_tracker.apply_file_derived_seed) -- export and
re-import round-trip the status, not just the raw data.

Sum Data's stage cells intentionally look different: literal "Done"/
"Open" text on light green/pink fills matching Excel's own built-in
"Good"/"Bad" conditional-formatting cell styles (see
_write_sum_data_stage_cell), not All Data's dark-fill "X"/"1" marks --
Sum Data is a status report, not a re-importable source sheet, so it's
free to read the way a human expects a status column to read.

Changes is three independent tables stacked in one sheet -- see
_write_changes_sheet -- State Revision Log (the state's own J-Changes
content, read-only reference data), Update History (this app's own
accumulated change_history.json, newest first), then Comments (this
app's own accumulated comments.json, newest first, including an
"Edited" column when a comment has been edited in place). State
Revision Log/Update History are read-only reference data that never
round-trips on re-import the way All Data's stage marks do; Comments is
free text a user typed directly into this app, same non-round-tripping
treatment.
"""

from __future__ import annotations

import re
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from core.excel_reader import STAGE_COUNT, live_sum_data_totals

_STAGE_MARKS = {"Done": "X", "Open": "1"}
_STAGE_FILLS = {
    "Done": PatternFill(start_color="FF2E7D32", end_color="FF2E7D32", fill_type="solid"),
    "Open": PatternFill(start_color="FFC62828", end_color="FFC62828", fill_type="solid"),
}
_STAGE_FONT = Font(color="FFFFFFFF", bold=True)
_STAGE_ALIGNMENT = Alignment(horizontal="center", vertical="center")

# Sum Data sheet's own required-stage cell style -- literal "Done"/"Open"
# text on a full-cell fill, matching Excel's own built-in "Good"/"Bad"
# conditional-formatting cell styles exactly (these are those styles'
# real fill/font colors), distinct from All Data's "X"/"1" dark-fill
# marks above -- Sum Data is meant to read like a status report, All
# Data like the source file's own raw markers.
_SUM_DATA_STAGE_LABELS = {"Done": "Done", "Open": "Open"}
_SUM_DATA_STAGE_FILLS = {
    "Done": PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid"),
    "Open": PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid"),
}
_SUM_DATA_STAGE_FONTS = {
    "Done": Font(color="FF006100"),
    "Open": Font(color="FF9C0006"),
}
_HEADER_FONT = Font(bold=True)
_BOLD_FONT = Font(bold=True)
_DESCRIPTION_ALIGNMENT = Alignment(wrap_text=True, vertical="top")

# Bottom totals row -- bold + a subtle tint, matching
# ui.pages.data_view_page's TOTALS_ROW_BACKGROUND exactly (#e4e8f0) so
# the export reads the same way the in-app tables do.
_TOTALS_FILL = PatternFill(start_color="FFE4E8F0", end_color="FFE4E8F0", fill_type="solid")
_TOTALS_FONT = Font(bold=True)

INDEX_COL = 1
DESCRIPTION_COL = 2
AGENCY_COL = 3
STAGE_FIRST_COL = 4
STAGE_LAST_COL = STAGE_FIRST_COL + STAGE_COUNT - 1

# Changes sheet -- three stacked tables sharing one column grid (see
# _write_changes_sheet). State Revision Log's long text (Synopsis),
# Update History's long text (Detail), and Comments' own free text
# deliberately all share column B, and Revision Log's three date
# columns / History's version+date columns / Comments' date share
# columns C-F, so one set of column widths below serves all three
# tables reasonably rather than needing per-section widths (which
# openpyxl has no way to express within a single column anyway).
REVISION_NUM_COL = 1
SYNOPSIS_COL = 2
AOR_SIGNATURE_COL = 3
SEOR_SIGNATURE_COL = 4
EFFECTIVE_DATE_COL = 5
HCAI_CONCURRENCE_COL = 6

UPDATE_NUM_COL = 1
DETAIL_COL = 2
SUMMARY_COL = 3
OLD_VERSION_COL = 4
NEW_VERSION_COL = 5
UPDATE_DATE_COL = 6

COMMENT_NUM_COL = 1
COMMENT_TEXT_COL = 2
COMMENT_DATE_COL = 3
COMMENT_EDITED_COL = 4

# Descriptions are 3 logical lines (component name, code citation, test
# description) joined with embedded \n. openpyxl can't auto-fit row height
# for wrapped text, so this is a fixed height sized for 3 lines at the
# default 11pt font (~15pt/line) plus a little padding.
_DESCRIPTION_ROW_HEIGHT = 50

_UNSAFE_FILENAME_CHARS = re.compile(r'[\\/:*?"<>|]')
_WHITESPACE = re.compile(r"\s+")


def default_filename(project_name: str, increment_name: str, version: int) -> str:
    """"{ProjectName}_{IncrementName}_v{version}.xlsx", with filesystem-
    unsafe characters (from either name -- project/increment names are
    free text the user typed or pulled from a workbook, not guaranteed
    filename-safe) stripped and whitespace collapsed to underscores.
    """

    def clean(text: str) -> str:
        stripped = _UNSAFE_FILENAME_CHARS.sub("", text)
        return _WHITESPACE.sub("_", stripped.strip())

    return f"{clean(project_name)}_{clean(increment_name)}_v{version}.xlsx"


def _write_header(ws: Worksheet, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT


def _stage_headers() -> list[str]:
    return [f"Stage {i}" for i in range(1, STAGE_COUNT + 1)]


def _set_common_column_widths(ws: Worksheet, extra_col_widths: dict[int, int]) -> None:
    ws.column_dimensions[get_column_letter(INDEX_COL)].width = 12
    ws.column_dimensions[get_column_letter(DESCRIPTION_COL)].width = 45
    ws.column_dimensions[get_column_letter(AGENCY_COL)].width = 20
    for col in range(STAGE_FIRST_COL, STAGE_LAST_COL + 1):
        ws.column_dimensions[get_column_letter(col)].width = 9
    for col, width in extra_col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def _apply_description_wrap(ws: Worksheet, excel_row: int, column: int = DESCRIPTION_COL) -> None:
    ws.cell(row=excel_row, column=column).alignment = _DESCRIPTION_ALIGNMENT
    ws.row_dimensions[excel_row].height = _DESCRIPTION_ROW_HEIGHT


def _write_stage_cell(ws: Worksheet, excel_row: int, stage: int, status: str | None) -> None:
    cell = ws.cell(row=excel_row, column=STAGE_FIRST_COL + stage - 1, value=_STAGE_MARKS.get(status, ""))
    fill = _STAGE_FILLS.get(status)
    if fill is not None:
        cell.fill = fill
        cell.font = _STAGE_FONT
        cell.alignment = _STAGE_ALIGNMENT


def _write_sum_data_stage_cell(ws: Worksheet, excel_row: int, stage: int, status: str) -> None:
    cell = ws.cell(
        row=excel_row, column=STAGE_FIRST_COL + stage - 1, value=_SUM_DATA_STAGE_LABELS[status]
    )
    cell.fill = _SUM_DATA_STAGE_FILLS[status]
    cell.font = _SUM_DATA_STAGE_FONTS[status]
    cell.alignment = _STAGE_ALIGNMENT


def _write_totals_row(ws: Worksheet, values: list[Any]) -> None:
    """Appends a bold, tinted summary row -- not a real item, so it's
    deliberately styled to stand apart from the data rows above it (see
    module docstring for where these totals come from and why All
    Data's are safe to compute once while Sum Data's are recomputed live
    on every export).
    """
    ws.append(values)
    for cell in ws[ws.max_row]:
        cell.font = _TOTALS_FONT
        cell.fill = _TOTALS_FILL
        if cell.column > DESCRIPTION_COL:
            cell.alignment = _STAGE_ALIGNMENT


def _write_all_data_sheet(ws: Worksheet, rows: list[dict[str, Any]], totals: dict[str, Any]) -> None:
    headers = ["Index", "Description", "Approval Agency"] + _stage_headers() + ["VCR", "SUM"]
    _write_header(ws, headers)

    for row_data in rows:
        required = set(row_data.get("required_stages", []))
        stage_status = row_data.get("stage_status", {})
        values = [
            row_data.get("index"),
            row_data.get("description") or "",
            row_data.get("approval_agency") or "",
        ]
        for stage in range(1, STAGE_COUNT + 1):
            if stage in required:
                values.append(_STAGE_MARKS.get(stage_status.get(stage), ""))
            else:
                # Matches the source file's own convention (and
                # core.excel_reader.build_all_data(), which already
                # stores 0 here) -- a genuine numeric 0, not blank.
                values.append(0)
        # Same "0, not blank" convention as non-required stage cells above.
        vcr = row_data.get("vcr")
        values.append(0 if vcr is None else vcr)
        values.append(row_data.get("sum"))
        ws.append(values)

        excel_row = ws.max_row
        _apply_description_wrap(ws, excel_row)
        for stage in required:
            _write_stage_cell(ws, excel_row, stage, stage_status.get(stage))
        for stage in range(1, STAGE_COUNT + 1):
            if stage not in required:
                ws.cell(row=excel_row, column=STAGE_FIRST_COL + stage - 1).alignment = _STAGE_ALIGNMENT

    totals_values = ["Totals", "", ""]
    for stage in range(1, STAGE_COUNT + 1):
        totals_values.append(totals.get(f"Stage {stage}", 0))
    totals_values.append(totals.get("VCR", 0))
    totals_values.append(totals.get("SUM", 0))
    _write_totals_row(ws, totals_values)

    vcr_col = STAGE_LAST_COL + 1
    sum_col = vcr_col + 1
    _set_common_column_widths(ws, {vcr_col: 9, sum_col: 9})
    ws.freeze_panes = ws.cell(row=2, column=STAGE_FIRST_COL).coordinate


def _write_sum_data_sheet(ws: Worksheet, rows: list[dict[str, Any]]) -> None:
    """Mirrors ui.pages.data_view_page's Sum Data tab exactly: live
    status.json-backed Done/Open per required stage (an unset required
    stage defaults to "Open" -- see that module's docstring), plus live
    Open/Done/Total/% Complete columns, rather than build_sum_data()'s
    own raw-file-derived status text.
    """
    headers = ["Index", "Description", "Approval Agency"] + _stage_headers() + [
        "VCR", "Open", "Done", "Total", "% Complete"
    ]
    _write_header(ws, headers)

    for row_data in rows:
        required = row_data.get("required_stages", [])
        stage_status = row_data.get("stage_status", {})
        live_status = {stage: stage_status.get(stage, "Open") for stage in required}
        done_count = sum(1 for status in live_status.values() if status == "Done")
        open_count = sum(1 for status in live_status.values() if status == "Open")
        total_count = len(required)
        pct = (done_count / total_count) if required else None

        values = [
            row_data.get("index"),
            row_data.get("description") or "",
            row_data.get("approval_agency") or "",
        ]
        values += ["" for _ in range(1, STAGE_COUNT + 1)]  # written properly below via _write_sum_data_stage_cell
        values.append(row_data.get("vcr"))
        values.append(open_count)
        values.append(done_count)
        values.append(total_count)
        values.append(pct)
        ws.append(values)

        excel_row = ws.max_row
        _apply_description_wrap(ws, excel_row)
        for stage, status in live_status.items():
            _write_sum_data_stage_cell(ws, excel_row, stage, status)

        pct_col = STAGE_LAST_COL + 5
        if pct is not None:
            ws.cell(row=excel_row, column=pct_col).number_format = "0%"

    vcr_col = STAGE_LAST_COL + 1
    open_col = vcr_col + 1
    done_col = vcr_col + 2
    total_col = vcr_col + 3
    pct_col = vcr_col + 4

    # Live totals -- recomputed from these SAME rows every export (see
    # core.excel_reader.live_sum_data_totals), never cached, so this
    # never goes stale relative to the individual rows just written above.
    totals = live_sum_data_totals(rows)
    totals_values = ["Totals", "", ""]
    for stage in range(1, STAGE_COUNT + 1):
        totals_values.append(totals["stage_open_counts"][stage])
    totals_values.append(totals["vcr_open_count"])
    totals_values.append(totals["open_total"])
    totals_values.append(totals["done_total"])
    totals_values.append(totals["grand_total"])
    totals_values.append(totals["pct_complete"])
    _write_totals_row(ws, totals_values)
    if totals["pct_complete"] is not None:
        ws.cell(row=ws.max_row, column=pct_col).number_format = "0%"

    _set_common_column_widths(ws, {vcr_col: 9, open_col: 9, done_col: 9, total_col: 9, pct_col: 12})
    ws.freeze_panes = ws.cell(row=2, column=STAGE_FIRST_COL).coordinate


REPORT_DESCRIPTION_COL = 3


def _write_report_sheet(ws: Worksheet, rows: list[dict[str, Any]]) -> None:
    headers = ["Approval Agency", "Index", "Description", "Total"]
    _write_header(ws, headers)

    for row_data in rows:
        is_grand_total = row_data.get("approval_agency") == "Grand Total"
        ws.append(
            [
                row_data.get("approval_agency") or "",
                row_data.get("index") or "",
                row_data.get("description") or "",
                row_data.get("total", 0),
            ]
        )
        _apply_description_wrap(ws, ws.max_row, column=REPORT_DESCRIPTION_COL)
        if is_grand_total:
            for cell in ws[ws.max_row]:
                cell.font = _BOLD_FONT

    ws.column_dimensions[get_column_letter(1)].width = 30
    ws.column_dimensions[get_column_letter(2)].width = 12
    ws.column_dimensions[get_column_letter(3)].width = 45
    ws.column_dimensions[get_column_letter(4)].width = 10


# ----------------------------------------------------------------------
# Changes sheet -- see _write_changes_sheet. The small phrasing helpers
# below are duplicated from ui/dialogs/review_dialog.py and
# ui/pages/data_view_page.py rather than imported (core/ doesn't import
# ui/ anywhere else in this codebase, and _format_total is already
# duplicated the same way in review_dialog.py/data_view_page.py) --
# same wording as those two so the export reads exactly like the
# Changes tab it mirrors.
# ----------------------------------------------------------------------
def _first_line(text: str | None) -> str:
    if not text:
        return ""
    return text.split("\n", 1)[0]


def _describe_stage_value(value) -> str:
    if value in (0, None, ""):
        return "not required"
    if isinstance(value, str) and value.strip().lower() == "x":
        return "required (X)"
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        n = int(value) if float(value).is_integer() else value
        return f"{n} test{'s' if n != 1 else ''} required"
    return str(value)


def _describe_change(field_label: str, old_value, new_value) -> str:
    if field_label.startswith("Stage "):
        return f"{field_label}: {_describe_stage_value(old_value)} → {_describe_stage_value(new_value)}"
    old_text = old_value if old_value not in (None, "") else "(blank)"
    new_text = new_value if new_value not in (None, "") else "(blank)"
    return f"{field_label}: {old_text!s} → {new_text!s}"


def _history_entry_summary(entry: dict) -> str:
    parts = []
    added = entry.get("added_items") or []
    removed = entry.get("removed_items") or []
    anomalies = entry.get("column_anomalies") or []
    value_changed = entry.get("value_changed_items") or []
    if added:
        parts.append(f"{len(added)} item{'s' if len(added) != 1 else ''} added")
    if removed:
        parts.append(f"{len(removed)} item{'s' if len(removed) != 1 else ''} removed")
    if value_changed:
        parts.append(f"{len(value_changed)} value{'s' if len(value_changed) != 1 else ''} changed")
    if anomalies:
        parts.append(f"{len(anomalies)} column anomal{'y' if len(anomalies) == 1 else 'ies'}")
    return ", ".join(parts) if parts else "No changes detected"


def _history_entry_detail_text(entry: dict) -> str:
    """Plain-text equivalent of what ui.pages.data_view_page's Changes
    tab shows when an Update History entry is expanded -- one \\n-joined
    string (a spreadsheet cell, not a widget tree) built from the SAME
    persisted fields, in the same order/grouping.
    """
    added = entry.get("added_items") or []
    removed = entry.get("removed_items") or []
    anomalies = entry.get("column_anomalies") or []
    value_changed = entry.get("value_changed_items") or []

    if not (added or removed or anomalies or value_changed):
        return "No changes detected -- this update replaced the file with an item-for-item identical upload."

    lines = []
    if added:
        lines.append(f"Added Items ({len(added)}):")
        lines.extend(f"  {item.get('index')} — {_first_line(item.get('description'))}" for item in added)
    if removed:
        lines.append(f"Removed Items ({len(removed)}):")
        lines.extend(f"  {item.get('index')} — {_first_line(item.get('description'))}" for item in removed)
    if value_changed:
        lines.append(f"Values Changed ({len(value_changed)}):")
        for item in value_changed:
            lines.append(f"  {item.get('index')} — {_first_line(item.get('description'))}")
            lines.extend(
                f"    {_describe_change(field_label, old_v, new_v)}"
                for field_label, (old_v, new_v) in (item.get("changes") or {}).items()
            )
    if anomalies:
        lines.append(f"Column Anomalies ({len(anomalies)}):")
        lines.extend(f"  {anomaly}" for anomaly in anomalies)
    return "\n".join(lines)


def _format_changes_date(value) -> str:
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return str(value)


def _wrapped_row_height(text: str) -> int:
    """Same per-line-height convention _DESCRIPTION_ROW_HEIGHT already
    uses (~15pt/line + padding) -- openpyxl can't auto-fit row height for
    wrapped text, so this estimates from the text's own embedded
    newlines (undercounts additional column-width-forced wraps, but is a
    much better estimate than a flat 3-line height: J-Changes' Synopsis
    text and Update History's Detail text both routinely run well past
    3 real lines -- see core.excel_reader.raw_changes_log).
    """
    line_count = (text or "").count("\n") + 1
    return max(_DESCRIPTION_ROW_HEIGHT, line_count * 15 + 10)


def _write_changes_sheet(
    ws: Worksheet, changes_log: list[dict], change_history: list[dict], comments: list[dict]
) -> None:
    """Three independent tables stacked in one sheet, each separated by
    a blank row -- State Revision Log (the state's own J-Changes
    content, see core.excel_reader.raw_changes_log), Update History
    (this app's accumulated change_history.json), then Comments (this
    app's accumulated comments.json) -- matching ui.pages.data_view_page's
    Changes tab section order, content, and (for Update History/
    Comments) newest-first ordering exactly.
    """
    section_row = 1
    ws.cell(row=section_row, column=1, value="State Revision Log").font = _BOLD_FONT

    header_row = section_row + 1
    revision_headers = ["Rev #", "Synopsis of Change", "AOR Signature", "SEOR Signature", "Effective Date", "HCAI Concurrence"]
    for col, label in enumerate(revision_headers, start=1):
        ws.cell(row=header_row, column=col, value=label).font = _HEADER_FONT

    row = header_row + 1
    for entry in changes_log:
        synopsis = entry.get("synopsis") or ""
        ws.cell(row=row, column=REVISION_NUM_COL, value=entry.get("revision_number"))
        ws.cell(row=row, column=SYNOPSIS_COL, value=synopsis).alignment = _DESCRIPTION_ALIGNMENT
        ws.cell(row=row, column=AOR_SIGNATURE_COL, value=_format_changes_date(entry.get("aor_signature_date")))
        ws.cell(row=row, column=SEOR_SIGNATURE_COL, value=_format_changes_date(entry.get("seor_signature_date")))
        ws.cell(row=row, column=EFFECTIVE_DATE_COL, value=_format_changes_date(entry.get("effective_date")))
        ws.cell(row=row, column=HCAI_CONCURRENCE_COL, value=entry.get("hcai_concurrence") or "")
        ws.row_dimensions[row].height = _wrapped_row_height(synopsis)
        row += 1
    if not changes_log:
        ws.cell(row=row, column=1, value="No revision log entries found in this version's J-Changes sheet.")
        row += 1

    row += 1  # blank separator row
    history_section_row = row
    ws.cell(row=history_section_row, column=1, value="Update History").font = _BOLD_FONT

    history_header_row = history_section_row + 1
    history_headers = ["Update #", "Detail", "Summary", "Old Version", "New Version", "Date"]
    for col, label in enumerate(history_headers, start=1):
        ws.cell(row=history_header_row, column=col, value=label).font = _HEADER_FONT

    row = history_header_row + 1
    # Newest first -- change_history is stored oldest-first (append
    # order, see core.project_store), same reversal
    # ui.pages.data_view_page's Changes tab applies for display.
    for i, entry in enumerate(reversed(change_history), start=1):
        detail = _history_entry_detail_text(entry)
        ws.cell(row=row, column=UPDATE_NUM_COL, value=i)
        ws.cell(row=row, column=DETAIL_COL, value=detail).alignment = _DESCRIPTION_ALIGNMENT
        ws.cell(row=row, column=SUMMARY_COL, value=_history_entry_summary(entry))
        ws.cell(row=row, column=OLD_VERSION_COL, value=entry.get("old_version"))
        ws.cell(row=row, column=NEW_VERSION_COL, value=entry.get("new_version"))
        ws.cell(row=row, column=UPDATE_DATE_COL, value=(entry.get("timestamp") or "").split("T", 1)[0])
        ws.row_dimensions[row].height = _wrapped_row_height(detail)
        row += 1
    if not change_history:
        ws.cell(row=row, column=1, value="No updates confirmed yet for this increment.")
        row += 1

    row += 1  # blank separator row
    comments_section_row = row
    ws.cell(row=comments_section_row, column=1, value="Comments").font = _BOLD_FONT

    comments_header_row = comments_section_row + 1
    comments_headers = ["#", "Comment", "Date", "Edited"]
    for col, label in enumerate(comments_headers, start=1):
        ws.cell(row=comments_header_row, column=col, value=label).font = _HEADER_FONT

    row = comments_header_row + 1
    # Newest first, same reversal/convention as Update History above --
    # comments.json is also stored oldest-first (append order).
    for i, entry in enumerate(reversed(comments), start=1):
        text = entry.get("text") or ""
        edited_timestamp = entry.get("edited_timestamp")
        ws.cell(row=row, column=COMMENT_NUM_COL, value=i)
        ws.cell(row=row, column=COMMENT_TEXT_COL, value=text).alignment = _DESCRIPTION_ALIGNMENT
        ws.cell(row=row, column=COMMENT_DATE_COL, value=(entry.get("timestamp") or "").split("T", 1)[0])
        ws.cell(row=row, column=COMMENT_EDITED_COL, value=edited_timestamp.split("T", 1)[0] if edited_timestamp else "")
        ws.row_dimensions[row].height = _wrapped_row_height(text)
        row += 1
    if not comments:
        ws.cell(row=row, column=1, value="No comments yet for this increment.")

    ws.column_dimensions[get_column_letter(1)].width = 10
    ws.column_dimensions[get_column_letter(2)].width = 70
    ws.column_dimensions[get_column_letter(3)].width = 45
    ws.column_dimensions[get_column_letter(4)].width = 14
    ws.column_dimensions[get_column_letter(5)].width = 14
    ws.column_dimensions[get_column_letter(6)].width = 22


def export_increment(increment: Any, path: str) -> None:
    """Writes increment (a ui.mock_data.Increment, with all_data/
    sum_data/report/changes_log/change_history already populated by
    MockDataStore.get_increment_for_display) to a new .xlsx workbook at
    path with four sheets, in order: All Data, Sum Data, Report, Changes.
    """
    wb = openpyxl.Workbook()
    ws_all_data = wb.active
    ws_all_data.title = "All Data"
    _write_all_data_sheet(ws_all_data, increment.all_data, increment.all_data_totals)

    ws_sum_data = wb.create_sheet("Sum Data")
    _write_sum_data_sheet(ws_sum_data, increment.sum_data)

    ws_report = wb.create_sheet("Report")
    _write_report_sheet(ws_report, increment.report)

    ws_changes = wb.create_sheet("Changes")
    _write_changes_sheet(ws_changes, increment.changes_log, increment.change_history, increment.comments)

    wb.save(path)

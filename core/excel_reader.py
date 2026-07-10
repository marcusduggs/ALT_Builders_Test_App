"""
TIO workbook normalization engine.

Reconstructs the three "red" output sheets (All Data, Sum Data, Report) from
the "blue" input sheets, in memory, instead of trusting whatever is cached in
those sheets on disk. This module was written by reverse-engineering the
formulas in tests/fixtures/sample_increment.xlsm (opened with
openpyxl(data_only=False) so formulas were visible, not just cached values).
Everything below is what was actually observed in that file -- not a guess at
what the template "should" do.

======================================================================
WORKBOOK SHAPE
======================================================================
Blue (input) sheets, in tab order: A-Project Info, B-Tests,
C-On-Site Special Inspections, D-Off-Site Special Inspections,
F-Cons Verif, G-IOR, E-Required Compliance Forms, H-Plan Review,
I-TIO Program Approval, J-Changes.

Of those, only THREE actually feed "All Data": B-Tests,
C-On-Site Special Inspections, D-Off-Site Special Inspections. All three
share one schema, referred to below as a "grid sheet":

    col D (4)          Index               e.g. "B-F1"
    col E..AT (5..46)  Stage 1 .. Stage 42  one column per stage, values are
                                            blank, 0, 1 ("Open"/required-not-
                                            done) or "X"/"x" ("Done")
    col AU (47)        Description (long text: code ref + item name)
    col AW (49)        OPAA No. / expiration date (carried through per-record
                        but NOT pulled into All Data by the original formulas)
    col AX (50)        Approval Agency

F-Cons Verif ALSO feeds All Data, but it is a structurally different sheet
(a milestone/reference-number log, not a per-stage grid) and only two of its
columns end up in All Data:

    col D (4)   Reference Number (bare int, 1..47)
    col F (6)   Milestone description
    col V (22)  VCR status: "X"/"x" = Done, 1 = Open (required, not done),
                other text (e.g. "NA") passes through unchanged

The remaining five blue sheets (E-Required Compliance Forms, G-IOR,
H-Plan Review, I-TIO Program Approval, J-Changes) are sign-off /
compliance-form sheets with a completely different, narrow layout (checkbox
+ signature tables). Grepping every formula in All Data and Sum Data for
sheet references confirms NONE of them are pulled into All Data, Sum Data,
or Report. They are out of scope for this engine, same as "Inputs" (hidden
facility lookup) and "Revison History".

======================================================================
PARSING STRATEGY (identity-based, not position-based)
======================================================================
The state re-sends an "updated" file for the SAME increment sometimes, not
just brand-new increments -- and an update can add/remove rows or shuffle
row order, not just change values in place. A parser keyed to fixed row
numbers (or even a computed contiguous row range, which is what the first
version of this module did) breaks the moment that happens. So:

    Each blue-sheet row becomes a record because its IDENTITY cell is
    non-blank -- Index (col D) for the grid sheets, Reference Number
    (col D) for F-Cons Verif -- not because of a row number, or a range
    computed from the first/last non-blank identity cell. Every row below
    the header is checked independently. Rows can be inserted, deleted, or
    reordered in a new version of the file and every OTHER row's record is
    unaffected.

    Column layout is still position-based (stage 1 is still "whatever is 5
    columns right of Index", description is still col AU, etc.) -- that is
    a deliberate scope boundary, not an oversight. If a future file's
    column layout actually changes (a stage column added/removed/moved),
    this parser will silently misread it, which is exactly why
    core/structure_diff.py exists: it compares two versions of the same
    file and flags column-layout drift for a human to look at, rather than
    this module trying to guess the new layout and adapt on its own.

======================================================================
ALL DATA
======================================================================
Columns: B=Inc # (row2 label), C=Index, D=Description, E=Apprval Agency,
F..AU=Stage 1..42, AV=VCR, AW=SUM.

For B-Tests / C-On-Site / D-Off-Site, All Data literally drags the same
formula down a contiguous row range per sheet:

    All Data row 3   = 'B-Tests'!row 8    (offset -5)
    All Data row 163 = 'C-On-Site ...'!row 8   (offset +155)
    All Data row 259 = 'D-Off-Site ...'!row 9  (offset +250)

Per row: C=srcD (Index), D=srcAU (Description), E=srcAX (Approval Agency),
F..AU = src E..AT (Stage 1..42, each dest column = src column + 1),
AW = SUM(F:AV) of that row.

IMPORTANT: the ORIGINAL TEMPLATE's formulas do a *contiguous row copy*, not
a "skip blank rows" copy -- category/sub-category header rows inside
B-Tests/C-.../D-... (e.g. B-Tests row 16, which only has "STRUCTURAL
TESTS"/"Concrete" text in cols A-C and nothing in D) fall inside the copied
formula range and come through as all-zero rows in the real All Data,
because Excel resolves a formula referencing a blank cell to 0.

This engine does NOT reproduce that quirk (as of the identity-based parser
rewrite -- see "PARSING STRATEGY" above): a blank-Index row is simply
skipped, not carried through as a zero-value record. Nothing downstream
ever used those placeholder rows (Sum Data and Report both filter on a
real Index anyway), and once row position stopped being load-bearing for
this parser there was no other reason to manufacture them. This changes
All Data's row *count* relative to the earlier version of this module (and
relative to the real file's All Data sheet) but not any Index's field
values, and tests/test_normalization.py's keyed comparison already ignored
blank-Index rows on both sides, so validation results are unaffected.

The row numbers above are quoted only as evidence of what the *original
template's formulas* do on this one fixture -- the parser in this module
does not hardcode them, or any row range at all. See "PARSING STRATEGY"
below for how records are actually located.

*** KNOWN BUG IN THE GROUND-TRUTH FILE (not replicated) ***
D-Off-Site's real block starts at source row 9, not row 8. Row 8 there
(Index "D-C1", description "Concrete ... Placement of concrete") is a
completely valid record -- it just has zero stages checked, same as several
other valid rows that DO appear in All Data (e.g. B-F3). It is simply
missing from the ground-truth All Data/Sum Data/Report: grepping the whole
workbook for the literal string "D-C1" outside of D-Off-Site itself returns
nothing. This engine's dynamic boundary detection intentionally does NOT
reproduce this drop -- for a normalization engine, silently discarding a
real record whenever the pattern happens to match B-tests' fill-down-and-
forget artifact is worse than a one-row mismatch against this one fixture.
tests/test_normalization.py calls this out explicitly rather than silently
matching it.

F-Cons Verif does NOT get a contiguous copy. Its mapping only includes rows
where the Reference Number (col D) is non-blank (43 rows 7..49, then a
second block rows 86..89 -- rows 50..85 are stray leftover "N" markers in
column V with nothing else, and are skipped entirely, not copied as blank
rows). This engine reproduces that by filtering to non-blank Reference
Number, which is the general form of what the fixture's row selection does.
All Data's Index for these rows is the static string "F-{reference_number}"
(e.g. "F-1") -- confirmed to hold for all 47 rows even though the literal
cells in the fixture are pasted values, not live formulas. Approval Agency
and all 42 Stage columns are left genuinely blank (None, not 0) for these
rows, because the original template never put a formula there at all.

======================================================================
SUM DATA
======================================================================
Columns: B..E mirror All Data (Inc #/Index/Description/Apprval Agency),
F..AU = Stage 1..42, AV = VCR, AW unused/blank, AX = COUNTIF(*Open*),
AY = COUNTIF(*Done*), AZ = AX+AY, BA = AY/AZ (raw "#DIV/0!" string when
AZ == 0, reproduced verbatim since that's what Excel would show).
Row 1 = VLOOKUP(stage_no, 'A-Project Info'!$D25:$F66, 3) -> stage name.

Sum Data is a FILTERED, reordered-by-nothing (i.e. original top-to-bottom
order preserved) subset of All Data:

    included if:  source_sheet == 'F-Cons Verif'
                  OR has at least one required stage (same definition
                  required_stages_by_index() uses: a stage value not in
                  (0, None, ""))

All 47 F-Cons Verif rows are included unconditionally.

*** BUG FIXED: inclusion used to require SUM > 0 AND a non-blank Approval
Agency instead ***
That was empirically a zero-mismatch match against this one fixture's
344 named All Data rows, which is exactly how it slipped through: the
one historical ground-truth file's real Sum Data sheet happens to ALSO
exclude every row that rule excludes, so byte-for-byte fidelity against
it never exercised the cases where the rule is wrong. Both conditions
were wrong relative to required_stages_by_index()'s definition, not just
stale-by-drift:
  - SUM (BlueSheetRecord.sum_value) only tallies NUMERIC stage values --
    it mirrors Excel's own SUM(), which ignores text cells -- so an item
    whose required stages are ALL "X"/Done (no numeric "1"/Open cells at
    all) has SUM == 0 and was silently excluded outright, even with a
    real Approval Agency. Confirmed on tests/fixtures/demo_after.xlsm's
    B-C18/B-C19/B-C20 (added for a later session's video-walkthrough
    demo, well after this fixture and its ground truth existed) -- all
    three have real Approval Agencies and real required stages, all
    marked "X", so SUM == 0 for all three and they vanished from Sum
    Data entirely on that upload.
  - A blank Approval Agency excluded a row regardless of having real
    required stages needing tracking. Confirmed on this fixture's own
    C-F4: required stages {10, 11, 12, 17, 18, 20} (all "1"/Open, so
    SUM == 6), Approval Agency blank -- excluded despite six real
    required stages.
This means the engine's Sum Data row SET can now differ from the
ground-truth fixture's real Sum Data sheet (C-F4 now correctly appears
in the engine's output; it never appears in the ground truth) --
tests/test_normalization.py documents this as a fifth named cause
alongside the pre-existing drift/bug causes below, not silently ignored.

Per-cell status transform (applied to each of the 42 stage columns + VCR):
    "x" / "X"        -> "Done"
    1                -> "Open"
    0 / None / ""    -> None (blank cell)
    anything else    -> passed through unchanged (e.g. "NA")

*** SUM DATA IS STALE RELATIVE TO THE BLUE SHEETS IN THIS FIXTURE ***
Unlike All Data, Sum Data has NO formulas in its Description/Approval
Agency/status columns (confirmed: 0 formula cells outside row 1 and the
per-row AX:BA completion math) -- it is hand-typed/pasted, and in this
fixture it has drifted from the live blue-sheet source in two distinct
ways, both confirmed by cross-checking against All Data (which, being
formula-driven, still agrees with the live blue sheets):
  1. Three Description cells contain "Done" spliced into the middle of a
     word -- "Excavation"->"EDonecavation", "Admixtures"->"AdmiDonetures",
     "Mix Design"->"MiDone Design". Pattern: the letter "x" in each word,
     replaced by "Done". Almost certainly a global, unanchored
     Find-and-Replace ("x" -> "Done") run directly on Sum Data at some
     point, that matched inside description text as a side effect (rather
     than only whole-cell "x" stage markers).
  2. One Approval Agency cell (row for C-FP3) is truncated to "CTS:" where
     the live source (and All Data) has the full
     "CTS: Jim Hollingsworth, Brad Thomas, ...". Independently-typed value
     that was never updated after the agency text was expanded upstream.
  3. Eight F-Cons Verif rows (F-5, F-6, F-7, F-9, F-10, F-11, F-12, F-20)
     show VCR="Done" in Sum Data, while F-Cons Verif's live V column (and
     All Data, which reads it live) has V=1 ("Open", not yet done). These
     items were reopened/reverted upstream after Sum Data was last
     hand-updated.
This engine reads only the blue sheets, so it reproduces the LIVE values,
not Sum Data's stale ones -- which is the whole point of the exercise
(the module-level task this was built for explicitly assumes red sheets
may be unreliable). tests/test_normalization.py surfaces all of these as
named, explained mismatches rather than silently matching the stale text.

======================================================================
REPORT
======================================================================
No formulas anywhere in the sheet (confirmed: 0 formula cells, and no
pivot-cache XML in the .xlsm package either) -- it is static/pasted values,
laid out like a two-level pivot table:

    Group by Approval Agency (outer), then Index (inner), of every All Data
    record whose Index is truthy (i.e. excludes the same blank-Index rows
    that never made it into Sum Data either).
    Total = SUM (All Data's AW column) per Index. (Every Index is unique in
    this fixture, but summing on collision is the correct general rule for
    a "Sum of SUM" pivot.)
    Sort: Approval Agency ascending (Python's default string sort matches
    the fixture for 10 of 11 groups); blank/falsy Approval Agency sorts
    LAST (all the F-* / F-Cons Verif rows). Within a group, Index ascending
    as plain text.
    Agency label is only shown on the first row of each group (blank on
    continuation rows), matching the fixture's display.
    Grand Total row at the bottom.

    One tie-break the fixture gets "wrong" vs. plain string sort: two
    Approval Agency values that differ only by a trailing space before an
    embedded newline ("Fenagh, \\nCTS" vs "Fenagh,\\nCTS") appear in the
    opposite order from a codepoint-by-codepoint sort. This is almost
    certainly Excel's locale-aware collation differing from Python's raw
    string comparison on a single whitespace-only tie -- documented here,
    not silently special-cased, since it is a single tie in one fixture and
    "fixing" it would mean guessing at Excel's collation rules.

Report inclusion is a THIRD, distinct filter from Sum Data's -- confirmed
empirically (58 grid rows with a real Approval Agency, regardless of SUM,
plus all 47 F-Cons Verif rows = exactly 105, matching the fixture's named
row count exactly):

    included if:  source_sheet == 'F-Cons Verif'
                  OR Approval Agency is non-blank
    (no SUM > 0 requirement, unlike Sum Data)

*** TWO UNEXPLAINED ANOMALIES IN THE GROUND-TRUTH FILE (not replicated) ***
1. The real Report has a trailing row, right before "Grand Total", reading
   Approval Agency/Index/Description = "(blank)"/"(blank)"/"(blank)" with a
   Total of 295 -- which is EXACTLY equal to the sum of every real row
   above it (the 105 named-index rows also sum to 295). Grand Total is then
   590, i.e. double-counted. There is no live formula, no pivot cache, and
   no blank-Index All Data row (real or synthetic) whose total is anywhere
   close to 295 -- the true blank-Index All Data rows all sum to 0.
2. One row (Approval Agency "CTS", Index "D-C7") shows Total = 20 in the
   real Report, but the underlying All Data row for D-C7 has SUM = 10 (one
   row only, no duplicate D-C7 anywhere else in the workbook) -- again
   exactly double.
Both anomalies are the same shape (an exact 2x of a value that is directly
verifiable elsewhere in the same workbook), which points at a manual
copy/paste duplication in this specific increment's Report tab rather than
a rule this engine should encode. This engine's reconstructed Report omits
both, and reports the arithmetically correct Grand Total (sum of the named
rows, no duplication). tests/test_normalization.py flags both explicitly
rather than matching them.

Separately, one Description in the real Report reads "Not Used" (row for
F-45) where the live F-Cons Verif source (and this engine) has "NOT USED"
verbatim -- consistent with Report being hand-typed/pasted rather than
formula-driven (see the Sum Data staleness note above; Report appears to
share that property for at least this one cell).
"""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import Any

import openpyxl
import pandas as pd

STAGE_COUNT = 42

GRID_SHEETS = [
    "B-Tests",
    "C-On-Site Special Inspections",
    "D-Off-Site Special Inspections",
]
MILESTONE_SHEET = "F-Cons Verif"
IGNORED_BLUE_SHEETS = [
    "E-Required Compliance Forms",
    "G-IOR",
    "H-Plan Review",
    "I-TIO Program Approval",
    "J-Changes",  # not ignored entirely -- see raw_changes_log() below,
    # which parses it independently for read-only display. Still excluded
    # from parse_workbook/normalize_workbook: it doesn't feed All Data,
    # Sum Data, or Report, so it stays out of that pipeline.
]

# --- grid sheet (B-Tests / C-On-Site / D-Off-Site) column layout ---
GRID_INDEX_COL = 4       # D
GRID_STAGE_FIRST_COL = 5  # E  -> stage 1
GRID_STAGE_LAST_COL = 46  # AT -> stage 42
GRID_DESC_COL = 47        # AU
GRID_OPAA_COL = 49        # AW
GRID_AGENCY_COL = 50      # AX
GRID_HEADER_LABEL = "index #"

# --- F-Cons Verif (milestone sheet) column layout ---
MILESTONE_REF_COL = 4      # D
MILESTONE_DESC_COL = 6     # F
MILESTONE_VCR_COL = 22     # V
MILESTONE_HEADER_LABEL = "reference"

# --- A-Project Info fixed cells ---
PROJECT_INFO_FACILITY_NUM_CELL = "D12"
PROJECT_INFO_FACILITY_NAME_CELL = "F12"
PROJECT_INFO_PROJECT_NUM_CELL = "I12"
PROJECT_INFO_RECORD_NAME_CELL = "G15"
PROJECT_INFO_SHEET = "A-Project Info"
RECORD_NAME_LABEL = "record name"
RECORD_NAME_SEARCH_ROWS = range(1, 30)
RECORD_NAME_SEARCH_COLS = range(1, 20)
STAGE_NAME_FIRST_ROW = 25
STAGE_NAME_LAST_ROW = 66
STAGE_NAME_NUM_COL = 4  # D
STAGE_NAME_NAME_COL = 6  # F

# --- J-Changes (state's revision log) column layout -- reverse-engineered
# directly from tests/fixtures/sample_increment.xlsm's J-Changes sheet, same
# rigor as every other sheet above. A flat row-per-revision table, data rows
# 7+, NOT the Index/Stage grid the other blue sheets use. Column header text
# for D..H lives in row 6; column I's header is actually cell I4, merged
# I4:I6 (spans the header band, not inline with row 6's other headers).
CHANGES_SHEET = "J-Changes"
CHANGES_HEADER_ROW = 6
CHANGES_REVISION_COL = 4        # D -- "REVISION NUMBER", int
CHANGES_SYNOPSIS_COL = 5        # E -- "SYNOPSIS OF CHANGE", long text
CHANGES_AOR_SIGNATURE_COL = 6   # F -- AOR Signature (Initial/date), a date value (not split initial/date)
CHANGES_SEOR_SIGNATURE_COL = 7  # G -- SEOR Signature (Initial/date), same as F
CHANGES_EFFECTIVE_DATE_COL = 8  # H -- "DATE of Effective Change"
CHANGES_HCAI_COL = 9            # I -- HCAI FDD Concurrence (Initial/date), free text "name\ndate"; often blank (not yet concurred)


@dataclass
class BlueSheetRecord:
    """One flattened line item, sourced from a single blue-sheet row."""

    source_sheet: str
    source_row: int
    index: str | None
    description: Any
    approval_agency: Any
    opaa_info: Any
    stage_values: dict[int, Any] = field(default_factory=dict)
    vcr: Any = None  # only meaningful for F-Cons Verif records
    # Fill-color-derived status ("Done"/"Open"), keyed by stage number --
    # only populated for grid-sheet records (see fill_status_by_index()
    # docstring for why this is scoped the same way required_stages_by_index()
    # scopes VCR out: a deliberate carry-over of that same boundary).
    stage_fill_status: dict[int, str] = field(default_factory=dict)

    @property
    def sum_value(self) -> float:
        """Replicates All Data's AW = SUM(F:AV): numeric cells only."""
        total = 0
        for v in self.stage_values.values():
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                total += v
        if isinstance(self.vcr, (int, float)) and not isinstance(self.vcr, bool):
            total += self.vcr
        return total


def _blank_to_none(value):
    return None if value in (0, "", None) else value


def _fill_color_status(cell) -> str | None:
    """Classifies a cell's fill/background color as "Done" (green),
    "Open" (red), or None (no color signal) -- see
    reys_fill_color_convention in project memory for why this is trusted
    as Rey's team's own status mark, not derived/guessed from anything.

    Only resolves plain RGB and indexed-palette fill colors. Theme-based
    fill colors (Excel's color-picker top row, which reference the
    workbook's theme + a tint rather than a fixed RGB triple) are
    deliberately NOT resolved -- doing so correctly requires parsing the
    workbook's theme XML, which this engine doesn't do. A themed fill
    color is treated the same as no color at all: falls back to
    needs_status, same as before this feature existed, rather than
    guessing.
    """
    fill = cell.fill
    if fill is None or fill.patternType is None:
        return None
    color = fill.fgColor
    if color is None:
        return None

    rgb = None
    if color.type == "rgb" and isinstance(color.rgb, str) and len(color.rgb) == 8:
        rgb = color.rgb
    elif color.type == "indexed":
        from openpyxl.styles.colors import COLOR_INDEX

        try:
            rgb = COLOR_INDEX[color.indexed]
        except (IndexError, TypeError):
            return None
    if rgb is None:
        return None

    try:
        r, g, b = int(rgb[2:4], 16), int(rgb[4:6], 16), int(rgb[6:8], 16)
    except ValueError:
        return None

    margin = 30
    if g > r + margin and g > b + margin:
        return "Done"
    if r > g + margin and r > b + margin:
        return "Open"
    return None


def _find_header_row(ws, col: int, label: str, search_rows=range(1, 30)) -> int:
    for r in search_rows:
        v = ws.cell(row=r, column=col).value
        if isinstance(v, str) and v.strip().lower().startswith(label):
            return r
    raise ValueError(f"Could not find header row with '{label}' in column {col} of {ws.title!r}")


def parse_grid_sheet(ws) -> list[BlueSheetRecord]:
    """Shared parser for B-Tests / C-On-Site / D-Off-Site (same schema).

    Identity-based: a row becomes a record because its Index cell (col D) is
    non-blank, full stop -- not because it falls between a computed start
    row and end row. Every row below the header is checked independently,
    so rows can be inserted, deleted, or reordered in a future version of
    the same increment without breaking this parser, as long as the column
    layout (which stage lives in which column, description, agency) stays
    put -- column-layout changes are a job for core/structure_diff.py to
    *detect*, not for this function to silently adapt to.

    Category/sub-category header rows (blank Index, e.g. "STRUCTURAL TESTS"
    text in cols A-C with nothing in D) are simply skipped -- they are not
    carried through as blank placeholder records. (An earlier version of
    this parser walked a fixed contiguous row range and therefore included
    those rows as all-zero records, to mirror a quirk of how the original
    template's formulas were dragged down. That quirk was a row-position
    artifact; once row position stopped being load-bearing there was
    nothing left to replicate it for -- All Data's downstream consumers
    never used those rows anyway. See the module docstring's "ALL DATA"
    section for the historical detail.)
    """
    header_row = _find_header_row(ws, GRID_INDEX_COL, GRID_HEADER_LABEL)

    records = []
    for r in range(header_row + 1, ws.max_row + 1):
        index = _blank_to_none(ws.cell(row=r, column=GRID_INDEX_COL).value)
        if index is None:
            continue
        stage_values = {}
        stage_fill_status = {}
        for stage in range(1, STAGE_COUNT + 1):
            cell = ws.cell(row=r, column=GRID_STAGE_FIRST_COL + stage - 1)
            stage_values[stage] = cell.value
            fill_status = _fill_color_status(cell)
            if fill_status is not None:
                stage_fill_status[stage] = fill_status
        records.append(
            BlueSheetRecord(
                source_sheet=ws.title,
                source_row=r,
                index=index,
                description=_blank_to_none(ws.cell(row=r, column=GRID_DESC_COL).value),
                approval_agency=_blank_to_none(ws.cell(row=r, column=GRID_AGENCY_COL).value),
                opaa_info=_blank_to_none(ws.cell(row=r, column=GRID_OPAA_COL).value),
                stage_values=stage_values,
                stage_fill_status=stage_fill_status,
            )
        )
    return records


def parse_milestone_sheet(ws) -> list[BlueSheetRecord]:
    """Parser for F-Cons Verif: a milestone log, not a per-stage grid.

    Already identity-based: every row below the header is checked
    independently, and a row becomes a record because its Reference Number
    cell (col D) is non-blank -- not because of a computed row range. Rows
    can be inserted, deleted, or reordered without breaking this. (See
    module docstring -- the fixture skips 36 stray rows this way, not by
    copying a contiguous range.)
    """
    header_row = _find_header_row(ws, MILESTONE_REF_COL, MILESTONE_HEADER_LABEL)
    records = []
    for r in range(header_row + 2, ws.max_row + 1):
        ref = ws.cell(row=r, column=MILESTONE_REF_COL).value
        if not isinstance(ref, (int, float)) or isinstance(ref, bool):
            continue
        records.append(
            BlueSheetRecord(
                source_sheet=ws.title,
                source_row=r,
                index=f"F-{int(ref)}",
                description=ws.cell(row=r, column=MILESTONE_DESC_COL).value,
                approval_agency=None,
                opaa_info=None,
                stage_values={stage: None for stage in range(1, STAGE_COUNT + 1)},
                vcr=ws.cell(row=r, column=MILESTONE_VCR_COL).value,
            )
        )
    return records


def raw_changes_log(ws) -> list[dict[str, Any]]:
    """Parses J-Changes: the state's own revision log, read-only reference
    data that does NOT feed All Data/Sum Data/Report (see IGNORED_BLUE_SHEETS
    and the module docstring) -- kept as its own independent entry point
    rather than folded into parse_workbook/normalize_workbook, to preserve
    that scope boundary.

    Identity-based like parse_milestone_sheet: a row becomes an entry
    because its Revision Number cell (col D) is a real number, not because
    of a computed row range -- consistent with this module's general
    parsing strategy, even though in practice this sheet's revisions are
    always contiguous.

    HCAI Concurrence (col I) is frequently blank -- a revision awaiting
    sign-off, not a parsing error -- and is returned as None rather than
    dropped or defaulted to an empty string.
    """
    records = []
    for r in range(CHANGES_HEADER_ROW + 1, ws.max_row + 1):
        revision = ws.cell(row=r, column=CHANGES_REVISION_COL).value
        if not isinstance(revision, (int, float)) or isinstance(revision, bool):
            continue
        records.append(
            {
                "revision_number": int(revision),
                "synopsis": _blank_to_none(ws.cell(row=r, column=CHANGES_SYNOPSIS_COL).value),
                "aor_signature_date": _blank_to_none(ws.cell(row=r, column=CHANGES_AOR_SIGNATURE_COL).value),
                "seor_signature_date": _blank_to_none(ws.cell(row=r, column=CHANGES_SEOR_SIGNATURE_COL).value),
                "effective_date": _blank_to_none(ws.cell(row=r, column=CHANGES_EFFECTIVE_DATE_COL).value),
                "hcai_concurrence": _blank_to_none(ws.cell(row=r, column=CHANGES_HCAI_COL).value),
            }
        )
    return records


def parse_project_info(ws) -> dict[str, Any]:
    """Facility/project header fields + the Stage No. -> Stage Name lookup."""
    stage_names: dict[int, Any] = {}
    for r in range(STAGE_NAME_FIRST_ROW, STAGE_NAME_LAST_ROW + 1):
        num = ws.cell(row=r, column=STAGE_NAME_NUM_COL).value
        if isinstance(num, (int, float)) and not isinstance(num, bool):
            stage_names[int(num)] = ws.cell(row=r, column=STAGE_NAME_NAME_COL).value

    return {
        "facility_number": ws[PROJECT_INFO_FACILITY_NUM_CELL].value,
        "facility_name": ws[PROJECT_INFO_FACILITY_NAME_CELL].value,
        "project_number": ws[PROJECT_INFO_PROJECT_NUM_CELL].value,
        "record_name": ws[PROJECT_INFO_RECORD_NAME_CELL].value,
        "stage_names": stage_names,
    }


def _find_record_name_label_cell(ws):
    for row in RECORD_NAME_SEARCH_ROWS:
        for col in RECORD_NAME_SEARCH_COLS:
            value = ws.cell(row=row, column=col).value
            if isinstance(value, str) and value.strip().lower().startswith(RECORD_NAME_LABEL):
                return row, col
    return None


def get_record_name(path: Any) -> str:
    """Extracts the increment's real identity -- A-Project Info's "Record
    Name (Scope of Project)" field -- by searching for that label's text
    rather than assuming a fixed cell address (the same reasoning
    _find_header_row applies to the blue sheets: a label can drift to a
    different cell across template revisions, but its text doesn't).

    This is the source of truth core.increment_matcher uses to decide
    whether an upload is a new increment or a new version of an existing
    one -- unlike parse_project_info()'s "record_name" (a tolerant,
    fixed-cell read used only for cosmetic labeling, e.g. All Data's
    "Inc #" column), a wrong or missing answer here would silently
    misfile an upload, so this raises instead of returning None/blank on
    anything unexpected.

    `path` may be a file path (str/Path) or an already-open Workbook (see
    open_workbook) -- same convention as compare_structure()/compare_values().

    The returned value is trimmed and has internal whitespace collapsed
    to single spaces (needed for reliable exact-match comparison later --
    see core.increment_matcher), but is NOT case-folded or otherwise
    mangled: it's still the real display name, used as-is for the
    increment's name in the app.
    """
    wb = open_workbook(path)
    ws = wb[PROJECT_INFO_SHEET]

    label_cell = _find_record_name_label_cell(ws)
    if label_cell is None:
        raise ValueError(
            f"Could not find the 'Record Name (Scope of Project)' label in {PROJECT_INFO_SHEET!r} of {path!r}"
        )
    label_row, label_col = label_cell

    for col in range(label_col + 1, max(RECORD_NAME_SEARCH_COLS) + 1):
        value = ws.cell(row=label_row, column=col).value
        if value not in (None, ""):
            return " ".join(str(value).split()).strip()

    raise ValueError(
        f"Found the 'Record Name (Scope of Project)' label in {PROJECT_INFO_SHEET!r} of {path!r}, "
        "but no value next to it"
    )


def open_workbook(path_or_workbook: Any) -> openpyxl.Workbook:
    """Loads a TIO workbook, or passes one through unchanged if it's
    already loaded.

    Real TIO workbooks are large enough (macro-enabled, many sheets) that
    a single load can take several seconds -- callers that need the same
    file for more than one operation (e.g. core/structure_diff.py diffing
    two versions, or ui/mock_data.py both diffing and reading records from
    the same file) should load it once with this function and pass the
    resulting Workbook to each function that needs it, rather than passing
    a path to each and paying for a reload every time.
    """
    if isinstance(path_or_workbook, openpyxl.Workbook):
        return path_or_workbook
    path = str(path_or_workbook)
    # data_only=True: blue sheets are manual data entry, but a couple of
    # A-Project Info header cells (facility #/name) are themselves formulas
    # against the hidden Inputs lookup sheet, so we want cached values, not
    # formula source, for everything read here.
    return openpyxl.load_workbook(path, data_only=True, keep_vba=path.lower().endswith(".xlsm"))


def parse_workbook(path: Any) -> dict[str, Any]:
    """Parses every relevant blue sheet out of a TIO workbook.

    `path` may be a file path (str/Path) or an already-open Workbook (see
    open_workbook) -- pass an already-open one when the caller needs to
    reuse it for something else too, to avoid loading the same file twice.

    Returns {"records": [BlueSheetRecord, ...], "project_info": {...}}.
    Record order is: B-Tests block, then C-On-Site block, then D-Off-Site
    block, then F-Cons Verif block -- the same order All Data lays them out
    in.
    """
    wb = open_workbook(path)

    records: list[BlueSheetRecord] = []
    for sheet_name in GRID_SHEETS:
        records.extend(parse_grid_sheet(wb[sheet_name]))
    records.extend(parse_milestone_sheet(wb[MILESTONE_SHEET]))

    project_info = parse_project_info(wb["A-Project Info"])

    return {"records": records, "project_info": project_info}


ALL_DATA_COLUMNS = (
    ["Inc #", "Index", "Description", "Apprval Agency"]
    + [f"Stage {i}" for i in range(1, STAGE_COUNT + 1)]
    + ["VCR", "SUM"]
)


def build_all_data(records: list[BlueSheetRecord], inc_label: Any) -> pd.DataFrame:
    """Reconstructs All Data from parsed blue-sheet records.

    Blank source cells replicate Excel's "formula referencing a blank cell
    evaluates to 0" behavior for grid-sheet columns (Index/Description/
    Approval Agency/stages). F-Cons Verif rows keep genuine blanks (None)
    for Approval Agency and stages, since the real template never wired a
    formula there at all -- see module docstring.
    """
    rows = []
    for rec in records:
        is_grid = rec.source_sheet in GRID_SHEETS
        zero_default = 0 if is_grid else None
        row = {
            "Inc #": inc_label,
            "Index": rec.index if rec.index is not None else zero_default,
            "Description": rec.description if rec.description is not None else zero_default,
            "Apprval Agency": rec.approval_agency if rec.approval_agency is not None else zero_default,
        }
        for stage in range(1, STAGE_COUNT + 1):
            v = rec.stage_values.get(stage)
            row[f"Stage {stage}"] = v if v is not None else zero_default
        row["VCR"] = rec.vcr if rec.vcr is not None else zero_default
        row["SUM"] = rec.sum_value
        rows.append(row)
    return pd.DataFrame(rows, columns=ALL_DATA_COLUMNS)


SUM_DATA_COLUMNS = (
    ["Inc #", "Index", "Description", "Apprval Agency"]
    + [f"Stage {i}" for i in range(1, STAGE_COUNT + 1)]
    + ["VCR", "Open", "Done", "Total", "% Complete"]
)


def _status(value: Any) -> Any:
    """All Data raw value -> Sum Data status text. See module docstring."""
    if isinstance(value, str) and value.strip().lower() == "x":
        return "Done"
    if value == 1 and not isinstance(value, bool):
        return "Open"
    if value in (0, None, ""):
        return None
    return value


def build_sum_data(records: list[BlueSheetRecord], inc_label: Any) -> pd.DataFrame:
    """Reconstructs Sum Data: a filtered, status-transformed view of the
    same records used for All Data. Filter rule and status transform are
    documented in the module docstring.

    Inclusion rule for grid-sheet records: at least one required stage --
    the exact same definition required_stages_by_index() uses (a stage
    value not in (0, None, "")) -- not rec.sum_value > 0 and a non-blank
    Approval Agency, which is what this used to check. See the module
    docstring's SUM DATA section, "*** BUG FIXED ***", for why that was
    wrong: sum_value only tallies numeric ("1"/Open) stage values (it
    mirrors Excel's own SUM(), which ignores text cells), so an item
    whose required stages are ALL "X"/Done had sum_value == 0 and was
    silently excluded outright -- and a blank Approval Agency excluded an
    item regardless of having real required stages (e.g. C-F4).
    """
    rows = []
    for rec in records:
        is_milestone = rec.source_sheet == MILESTONE_SHEET
        if not is_milestone:
            has_required_stage = any(v not in (0, None, "") for v in rec.stage_values.values())
            if not has_required_stage:
                continue

        stage_statuses = {
            stage: _status(rec.stage_values.get(stage)) for stage in range(1, STAGE_COUNT + 1)
        }
        vcr_status = _status(rec.vcr)

        open_count = sum(1 for v in list(stage_statuses.values()) + [vcr_status] if v == "Open")
        done_count = sum(1 for v in list(stage_statuses.values()) + [vcr_status] if v == "Done")
        total = open_count + done_count
        pct_complete = (done_count / total) if total else "#DIV/0!"

        row = {
            "Inc #": inc_label,
            "Index": rec.index,
            "Description": rec.description,
            "Apprval Agency": rec.approval_agency,
        }
        for stage in range(1, STAGE_COUNT + 1):
            row[f"Stage {stage}"] = stage_statuses[stage]
        row["VCR"] = vcr_status
        row["Open"] = open_count
        row["Done"] = done_count
        row["Total"] = total
        row["% Complete"] = pct_complete
        rows.append(row)
    return pd.DataFrame(rows, columns=SUM_DATA_COLUMNS)


REPORT_COLUMNS = ["Apprval Agency", "Index", "Description", "Total"]


def build_report(records: list[BlueSheetRecord]) -> pd.DataFrame:
    """Reconstructs Report: group by (Approval Agency, Index) over every
    record with a truthy Index, sorted Agency (blank last) then Index, with
    a Grand Total row. See module docstring for the sort tie-break caveat
    and the ground-truth "(blank)" duplicate-total row this deliberately
    does not replicate.
    """
    named = [
        rec
        for rec in records
        if rec.index and (rec.source_sheet == MILESTONE_SHEET or rec.approval_agency)
    ]

    groups: dict[tuple[Any, str], dict[str, Any]] = {}
    for rec in named:
        key = (rec.approval_agency, rec.index)
        if key not in groups:
            groups[key] = {"description": rec.description, "total": 0.0}
        groups[key]["total"] += rec.sum_value

    def sort_key(key):
        agency, index = key
        return (agency is None or agency == "", agency or "", index)

    ordered_keys = sorted(groups.keys(), key=sort_key)

    rows = []
    last_agency = object()  # sentinel, never equals a real agency
    for key in ordered_keys:
        agency, index = key
        g = groups[key]
        if agency == last_agency:
            display_agency = None
        else:
            # First row of a new group: show the label, or the literal
            # "(blank)" placeholder for a blank-agency group (matches the
            # ground-truth fixture's pivot-style rendering of null groups).
            display_agency = agency if agency else "(blank)"
        rows.append(
            {
                "Apprval Agency": display_agency,
                "Index": index,
                "Description": g["description"],
                "Total": g["total"],
            }
        )
        last_agency = agency

    grand_total = sum(g["total"] for g in groups.values())
    rows.append({"Apprval Agency": "Grand Total", "Index": None, "Description": None, "Total": grand_total})

    return pd.DataFrame(rows, columns=REPORT_COLUMNS)


def normalize_workbook(path: Any) -> dict[str, pd.DataFrame]:
    """Top-level entry point: blue sheets in, {all_data, sum_data, report}
    out. `path` may be a file path (str/Path) or an already-open Workbook
    -- see parse_workbook/open_workbook.
    """
    parsed = parse_workbook(path)
    records = parsed["records"]
    inc_label = parsed["project_info"]["record_name"]

    all_data_df = build_all_data(records, inc_label)
    sum_data_df = build_sum_data(records, inc_label)

    return {
        "all_data": all_data_df,
        "sum_data": sum_data_df,
        "report": build_report(records),
        "project_info": parsed["project_info"],
        "fill_status": fill_status_by_index(records),
        "raw_status": raw_status_by_index(records),
        "all_data_totals": all_data_totals(all_data_df),
        "sum_data_totals": sum_data_totals(sum_data_df),
    }


def _numeric_column_sum(series: pd.Series) -> float:
    """Excel's own SUM() semantics -- adds numeric cells, silently skips
    text ones (e.g. "X" markers). BlueSheetRecord.sum_value already
    applies this same rule per row; this applies it per column, matching
    the real source file's own bottom-row =SUM(colX3:colX_last) formulas
    exactly (confirmed against sample_increment.xlsm's cached values).

    Explicitly excludes NaN despite it technically satisfying
    isinstance(v, float): a Stage/VCR column mixing grid-sheet rows
    (0-filled) with F-Cons Verif rows (genuinely None -- see
    build_all_data()'s docstring) gets upcast by pandas to float64,
    coercing those Nones to NaN -- the exact same coercion
    required_stages_by_index() already has to guard against for a
    different column. Without this check, Python's sum() silently
    returns NaN for the WHOLE column the moment a single NaN is in it,
    since any arithmetic involving NaN propagates NaN.
    """
    return sum(v for v in series if isinstance(v, (int, float)) and not isinstance(v, bool) and not pd.isna(v))


def all_data_totals(all_data: pd.DataFrame) -> dict[str, float]:
    """All Data's bottom totals row: a plain numeric sum of every Stage/
    VCR/SUM column, matching the real file's own formulas exactly (see
    module usage in normalize_workbook()).

    Unlike Sum Data's totals (see sum_data_totals() and
    live_sum_data_totals() below), these never depend on live
    status.json state -- a stage's raw numeric value doesn't change when
    its Done/Open status is toggled -- so this single, file-derived
    computation is correct wherever it's used (in-app and export), with
    no live variant needed.
    """
    totals: dict[str, float] = {}
    for stage in range(1, STAGE_COUNT + 1):
        totals[f"Stage {stage}"] = _numeric_column_sum(all_data[f"Stage {stage}"])
    totals["VCR"] = _numeric_column_sum(all_data["VCR"])
    totals["SUM"] = _numeric_column_sum(all_data["SUM"])
    return totals


def sum_data_totals(sum_data: pd.DataFrame) -> dict[str, Any]:
    """Sum Data's bottom totals row, computed straight from the RAW
    file-derived Sum Data DataFrame: COUNTIF(colX,"*Open*") per Stage/VCR
    column, plus SUM() of the per-row Open/Done/Total columns, plus
    overall % Complete = Done total / Total total -- matching the real
    file's own bottom-row formulas exactly (see module usage in
    normalize_workbook(), and the module docstring... see this
    function's callers for the exact formulas this replicates).

    This reflects the FILE's own raw Done/Open status
    (build_sum_data()'s _status() classification), not any live
    status.json override -- correct immediately after a fresh upload
    (status.json is seeded directly from this same raw classification on
    first upload -- see core.status_tracker), but NOT what the app
    actually displays/exports after a manual status edit. See
    live_sum_data_totals() for the version that stays correct then --
    that is the one ui.pages.data_view_page and core.excel_export
    actually use; this one is normalize_workbook()'s own file-level
    artifact.
    """
    totals: dict[str, Any] = {}
    for stage in range(1, STAGE_COUNT + 1):
        totals[f"Stage {stage}"] = int((sum_data[f"Stage {stage}"] == "Open").sum())
    totals["VCR"] = int((sum_data["VCR"] == "Open").sum())
    open_total = int(sum_data["Open"].sum())
    done_total = int(sum_data["Done"].sum())
    grand_total = int(sum_data["Total"].sum())
    totals["open_total"] = open_total
    totals["done_total"] = done_total
    totals["grand_total"] = grand_total
    totals["pct_complete"] = (done_total / grand_total) if grand_total else None
    return totals


def live_sum_data_totals(rows: list[dict]) -> dict[str, Any]:
    """The Sum Data totals row actually rendered in-app and exported --
    computed from the SAME live (required_stages, status.json-backed
    stage_status) data every individual Sum Data row already renders
    from (ui.mock_data.Increment.sum_data), so it can never go stale
    relative to what's on screen after a manual status edit. Recomputed
    fresh by every caller (the Sum Data tab on every render, the Excel
    export on every export) -- never cached.

    Mirrors sum_data_totals()'s formulas conceptually (per-stage Open
    count, VCR Open count, Open/Done/Total sums, overall % Complete) but
    walking live row dicts instead of the static, file-derived
    DataFrame. Equal to sum_data_totals() immediately after a fresh
    upload, since status.json is seeded directly from the file's own raw
    status at that point -- see core.status_tracker.
    """
    stage_open_counts = {stage: 0 for stage in range(1, STAGE_COUNT + 1)}
    vcr_open_count = 0
    open_total = 0
    done_total = 0
    for row in rows:
        required = row.get("required_stages", [])
        stage_status = row.get("stage_status", {})
        for stage in required:
            status = stage_status.get(stage, "Open")
            if status == "Done":
                done_total += 1
            else:
                stage_open_counts[stage] += 1
                open_total += 1
        vcr_status = _status(row.get("vcr"))
        if vcr_status == "Done":
            done_total += 1
        elif vcr_status == "Open":
            vcr_open_count += 1
            open_total += 1
    grand_total = open_total + done_total
    return {
        "stage_open_counts": stage_open_counts,
        "vcr_open_count": vcr_open_count,
        "open_total": open_total,
        "done_total": done_total,
        "grand_total": grand_total,
        "pct_complete": (done_total / grand_total) if grand_total else None,
    }


def required_stages_by_index(all_data: pd.DataFrame) -> dict[str, list[int]]:
    """Which of the 42 numbered stages are actually required for each
    named Index in an All Data DataFrame -- i.e. have a real marker
    (numeric 1 or "X"/"x"), not blank/0. This is the per-stage
    granularity core/status_tracker.py's carry_forward_status() operates
    on: Sum Data's Done/Open status lives at the (Index, Stage) level,
    not one flag per Index, so only genuinely-required stages should ever
    need or carry a status.

    F-Cons Verif ("F-N") indexes always come back with an empty list --
    none of their 42 Stage columns are ever populated (see the module
    docstring's ALL DATA section: F-Cons Verif tracks completion via a
    single VCR flag, not the 42-stage grid). Per-stage status tracking is
    scoped to the grid sheets' Stage columns only; VCR is unaffected and
    untracked by this function, a deliberate scope boundary carried over
    from the task this was built for, not an oversight.

    Uses pd.isna() rather than an `is None` check: build_all_data() puts
    a genuine None (not 0) in F-Cons Verif's Stage columns (see that
    function's docstring), but a pandas DataFrame column holding a mix of
    None and numeric/string values coerces those Nones to float `NaN`,
    which compares unequal to everything including itself -- `NaN not in
    (0, None, "")` is True, so an `in`-based check alone would silently
    (and wrongly) count every F-Cons Verif stage as required.
    """
    result: dict[str, list[int]] = {}
    for _, row in all_data.iterrows():
        index = row["Index"]
        if not index:
            continue
        required = [
            stage
            for stage in range(1, STAGE_COUNT + 1)
            if not pd.isna(row[f"Stage {stage}"]) and row[f"Stage {stage}"] not in (0, "")
        ]
        result[index] = required
    return result


def fill_status_by_index(records: list[BlueSheetRecord]) -> dict[str, dict[int, str]]:
    """Rey's team's fill-color-derived status ("Done"/"Open"), by Index,
    for every grid-sheet stage cell where a color signal was found -- see
    BlueSheetRecord.stage_fill_status and reys_fill_color_convention in
    project memory. Same VCR/milestone-sheet scope boundary as
    required_stages_by_index(): F-Cons Verif records never populate
    stage_fill_status, so they simply don't appear here.
    """
    result: dict[str, dict[int, str]] = {}
    for rec in records:
        if not rec.index or not rec.stage_fill_status:
            continue
        result.setdefault(rec.index, {}).update(rec.stage_fill_status)
    return result


def raw_status_by_index(records: list[BlueSheetRecord]) -> dict[str, dict[int, str]]:
    """Rey's team's Done/Open status, by Index, derived directly from
    each stage cell's raw marker ("X"/"x" -> Done, 1 -> Open) -- the same
    per-cell transform Sum Data itself uses (_status()), just computed
    across every grid-sheet row instead of Sum Data's filtered subset.

    Trusting the raw marker this directly is a deliberate choice, made
    with the "SUM DATA IS STALE..." caution above already on the table:
    that finding showed the raw marker can, in principle, disagree with
    Rey's team's actual determination (a stage reopened after being
    marked, for instance). The choice made anyway was that every
    required stage should resolve to a real status the moment a file is
    uploaded, rather than sitting unset -- since a grid-sheet cell is
    only ever blank, "X", or "1" to begin with, there's no genuine
    third/unknown state in the source file for the app to preserve by
    holding off. See core.status_tracker.apply_file_derived_seed, which
    only uses this as a fallback behind fill_status_by_index -- an
    explicit color mark from Rey's team still wins if the two disagree.
    """
    result: dict[str, dict[int, str]] = {}
    for rec in records:
        if not rec.index:
            continue
        stages = {}
        for stage, value in rec.stage_values.items():
            status = _status(value)
            if status in ("Done", "Open"):
                stages[stage] = status
        if stages:
            result[rec.index] = stages
    return result

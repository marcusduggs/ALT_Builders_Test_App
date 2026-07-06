"""
Structural change detection between two versions of the same TIO increment.

The state sometimes re-sends an "updated" file for an increment the app has
already parsed -- not a new increment. core/excel_reader.py's identity-based
parser (keyed off each row's Index/Reference Number, not its row position)
already tolerates rows being inserted, deleted, or reordered. This module
covers the other half: detecting when something actually changed, so a
human (Rey) can decide what to do about it. It does not try to resolve
anything on its own -- that is a deliberate scope boundary, not a missing
feature. See core/excel_reader.py's "PARSING STRATEGY" docstring section
for the reasoning this builds on.

For each of the four sheets that feed All Data (B-Tests, C-On-Site,
D-Off-Site, F-Cons Verif), compare_structure() reports:

  - added_indexes:    Index/Reference values present in the new file but
                       not the old one (new test items, or a renamed Index
                       landing as "new").
  - removed_indexes:  present in the old file but missing from the new one
                       (an item was deleted upstream, OR an Index changed
                       and this is its old identity going stale).
  - unchanged_indexes: present in both, for context/counts.
  - column_anomalies: human-readable strings describing any header text
                       drift at a column position this engine relies on
                       (see "COLUMN ANOMALY DETECTION" below) -- e.g. a
                       stage column's header no longer reads "Stage N...",
                       or is blank/different in the new file.

Both index and column comparisons are done directly between old-vs-new
files at each fixed column position core/excel_reader.py reads from --
there's no "expected schema" baked in separately from that module, so if
the parser's column constants change, this comparison logic doesn't fall
out of sync.

======================================================================
COLUMN ANOMALY DETECTION
======================================================================
Row position is allowed to drift (that's the whole point of the
identity-based parser); column position is NOT -- core/excel_reader.py
reads "stage 1" from "5 columns right of the Index column", full stop. If
a real state file ever inserts, removes, or reorders columns, the parser
will silently misread data at the wrong column. This module's job is to
make that loud instead of silent.

Detection method: for the header row (located the same way
core/excel_reader.py locates it -- by label text, not a hardcoded row
number) in both the old and new file, compare the (whitespace-normalized,
case-insensitive) header text at every column position the parser reads:
Index, each of the 42 Stage columns, Description, OPAA, Approval Agency
for grid sheets; Reference Number, Description, VCR for F-Cons Verif.

    - If the header text differs between old and new at a given position,
      that's flagged -- this catches a column being renamed, or (more
      commonly) a column being inserted/removed somewhere in the row,
      which shifts every fixed-position column after it and changes what
      text shows up at each position, even though no single column was
      "renamed" per se.
    - Stage columns get one extra check: if the new file's header text at
      that position doesn't contain the word "stage" at all, that's
      flagged even if it happens to equal the old text (belt-and-suspenders
      against both files having already drifted from the true schema).
    - If the header row itself can't be located in either file (the
      anchor label is missing entirely), that's flagged and no
      column-by-column comparison is attempted for that sheet.

This will NOT catch every possible structural change (e.g. a column
silently swapping meaning with another column of the same header text), and
it is not meant to. It is meant to catch the changes that would otherwise
corrupt output silently, and it errs toward over-flagging (a false-positive
anomaly is a five-second glance for a human; a missed one is silent data
corruption).
"""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

from core.excel_reader import (
    GRID_AGENCY_COL,
    GRID_DESC_COL,
    GRID_HEADER_LABEL,
    GRID_INDEX_COL,
    GRID_OPAA_COL,
    GRID_SHEETS,
    GRID_STAGE_FIRST_COL,
    MILESTONE_DESC_COL,
    MILESTONE_HEADER_LABEL,
    MILESTONE_REF_COL,
    MILESTONE_SHEET,
    MILESTONE_VCR_COL,
    STAGE_COUNT,
    BlueSheetRecord,
    _find_header_row,
    open_workbook,
    parse_grid_sheet,
    parse_milestone_sheet,
)

SHEETS_TO_COMPARE = [*GRID_SHEETS, MILESTONE_SHEET]


@dataclass
class SheetDiff:
    sheet: str
    added_indexes: list[str] = field(default_factory=list)
    removed_indexes: list[str] = field(default_factory=list)
    unchanged_indexes: list[str] = field(default_factory=list)
    column_anomalies: list[str] = field(default_factory=list)

    @property
    def has_changes(self) -> bool:
        return bool(self.added_indexes or self.removed_indexes or self.column_anomalies)


def _load(path: Any) -> openpyxl.Workbook:
    return open_workbook(path)


def _records_for_sheet(wb: openpyxl.Workbook, sheet_name: str) -> list[BlueSheetRecord]:
    ws = wb[sheet_name]
    if sheet_name == MILESTONE_SHEET:
        return parse_milestone_sheet(ws)
    return parse_grid_sheet(ws)


def _index_diff(old_records: list[BlueSheetRecord], new_records: list[BlueSheetRecord]):
    old_set = {r.index for r in old_records if r.index}
    new_set = {r.index for r in new_records if r.index}
    added = sorted(new_set - old_set)
    removed = sorted(old_set - new_set)
    unchanged = sorted(old_set & new_set)
    return added, removed, unchanged


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split()).strip().lower()


def _grid_column_anomalies(ws_old, ws_new, sheet_name: str) -> list[str]:
    anomalies: list[str] = []

    try:
        header_old = _find_header_row(ws_old, GRID_INDEX_COL, GRID_HEADER_LABEL)
    except ValueError:
        anomalies.append(f"{sheet_name}: could not locate the 'Index #' header row in the OLD file")
        header_old = None
    try:
        header_new = _find_header_row(ws_new, GRID_INDEX_COL, GRID_HEADER_LABEL)
    except ValueError:
        anomalies.append(
            f"{sheet_name}: could not locate the 'Index #' header row in the NEW file -- "
            "the Index column may have moved or the header text changed"
        )
        header_new = None
    if header_old is None or header_new is None:
        return anomalies

    positions = [("Index", GRID_INDEX_COL)]
    positions += [(f"Stage {s}", GRID_STAGE_FIRST_COL + s - 1) for s in range(1, STAGE_COUNT + 1)]
    positions += [("Description", GRID_DESC_COL), ("OPAA", GRID_OPAA_COL), ("Approval Agency", GRID_AGENCY_COL)]

    for label, col in positions:
        old_text = _normalize_header(ws_old.cell(row=header_old, column=col).value)
        new_text = _normalize_header(ws_new.cell(row=header_new, column=col).value)
        letter = get_column_letter(col)

        text_changed = old_text != new_text
        stage_lost_marker = label.startswith("Stage") and "stage" not in new_text

        if not text_changed and not stage_lost_marker:
            continue

        if not new_text:
            note = "column appears to be MISSING/blank in the new file"
        elif not old_text:
            note = "new file has header text here where the old file had none -- possible inserted column"
        elif stage_lost_marker:
            note = f"header text no longer reads like a stage column: {new_text!r}"
        else:
            note = f"header text changed from {old_text!r} to {new_text!r} -- possible rename or column shift"

        anomalies.append(f"{sheet_name}: {label} (col {letter}): {note}")

    return anomalies


def _milestone_column_anomalies(ws_old, ws_new, sheet_name: str) -> list[str]:
    anomalies: list[str] = []

    try:
        header_old = _find_header_row(ws_old, MILESTONE_REF_COL, MILESTONE_HEADER_LABEL)
    except ValueError:
        anomalies.append(f"{sheet_name}: could not locate the 'Reference Number' header row in the OLD file")
        header_old = None
    try:
        header_new = _find_header_row(ws_new, MILESTONE_REF_COL, MILESTONE_HEADER_LABEL)
    except ValueError:
        anomalies.append(
            f"{sheet_name}: could not locate the 'Reference Number' header row in the NEW file -- "
            "the Reference Number column may have moved or the header text changed"
        )
        header_new = None
    if header_old is None or header_new is None:
        return anomalies

    positions = [
        ("Reference Number", MILESTONE_REF_COL),
        ("Description", MILESTONE_DESC_COL),
        ("VCR", MILESTONE_VCR_COL),
    ]
    for label, col in positions:
        old_text = _normalize_header(ws_old.cell(row=header_old, column=col).value)
        new_text = _normalize_header(ws_new.cell(row=header_new, column=col).value)
        if old_text == new_text:
            continue
        letter = get_column_letter(col)
        if not new_text:
            note = "column appears to be MISSING/blank in the new file"
        elif not old_text:
            note = "new file has header text here where the old file had none -- possible inserted column"
        else:
            note = f"header text changed from {old_text!r} to {new_text!r} -- possible rename or column shift"
        anomalies.append(f"{sheet_name}: {label} (col {letter}): {note}")

    return anomalies


def compare_structure(old_workbook_path: Any, new_workbook_path: Any) -> dict[str, SheetDiff]:
    """Compares two versions of the same TIO increment, sheet by sheet.

    Each argument may be a file path (str/Path) or an already-open
    Workbook (see core.excel_reader.open_workbook) -- pass an already-open
    one when the caller needs to reuse it for something else too, to avoid
    loading the same (potentially several-second-to-parse) file twice.

    Detection only -- never resolves, merges, or picks a "correct" version.
    Every field is meant to be read by a human before anything downstream
    trusts the new file.
    """
    wb_old = _load(old_workbook_path)
    wb_new = _load(new_workbook_path)

    results: dict[str, SheetDiff] = {}
    for sheet_name in SHEETS_TO_COMPARE:
        old_records = _records_for_sheet(wb_old, sheet_name)
        new_records = _records_for_sheet(wb_new, sheet_name)
        added, removed, unchanged = _index_diff(old_records, new_records)

        if sheet_name == MILESTONE_SHEET:
            anomalies = _milestone_column_anomalies(wb_old[sheet_name], wb_new[sheet_name], sheet_name)
        else:
            anomalies = _grid_column_anomalies(wb_old[sheet_name], wb_new[sheet_name], sheet_name)

        results[sheet_name] = SheetDiff(
            sheet=sheet_name,
            added_indexes=added,
            removed_indexes=removed,
            unchanged_indexes=unchanged,
            column_anomalies=anomalies,
        )

    return results


def has_any_changes(results: dict[str, SheetDiff]) -> bool:
    return any(diff.has_changes for diff in results.values())


if __name__ == "__main__":
    import sys

    if len(sys.argv) != 3:
        print("Usage: python -m core.structure_diff <old_workbook.xlsm> <new_workbook.xlsm>")
        raise SystemExit(2)

    results = compare_structure(sys.argv[1], sys.argv[2])
    for sheet_name, diff in results.items():
        print(f"--- {sheet_name} ---")
        print(f"  unchanged: {len(diff.unchanged_indexes)}")
        print(f"  added:     {diff.added_indexes}")
        print(f"  removed:   {diff.removed_indexes}")
        print(f"  column anomalies: {diff.column_anomalies or 'none'}")
    print()
    print("Changes detected -- human review required." if has_any_changes(results) else "No structural changes detected.")

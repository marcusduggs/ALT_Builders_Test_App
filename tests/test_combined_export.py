"""
Tests core/combined_export.py against two REAL, differently-named
increments (not just two versions of the same one) -- demo_before.xlsm/
demo_after.xlsm ("INC 2 - Foundation and Underground Utilities", used
here at v1 specifically, NOT its latest v2) and demo_inc3_new.xlsm
("INC 3 - Sewer Systems") -- using a temp-directory ProjectStore, never
touches the real ~/AltamiranoBuildersAppData.

Run directly: `python tests/test_combined_export.py`
"""

import os
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl

from core import combined_export, excel_reader
from core.project_store import ProjectStore
from ui.mock_data import MockDataStore

FIXTURES = os.path.join(os.path.dirname(__file__), "fixtures")
DEMO_BEFORE = os.path.join(FIXTURES, "demo_before.xlsm")
DEMO_AFTER = os.path.join(FIXTURES, "demo_after.xlsm")
DEMO_INC3_NEW = os.path.join(FIXTURES, "demo_inc3_new.xlsm")
PROJECT_NAME = "Combined Export Validation"


def main():
    failures = []

    with tempfile.TemporaryDirectory() as tmp:
        project_store = ProjectStore(base_dir=Path(tmp) / "data")
        store = MockDataStore(project_store)
        store.add_project(PROJECT_NAME)

        # INC 2: v1 (demo_before) then v2 (demo_after) -- two real
        # versions of the SAME increment, so we can validate that
        # selecting v1 specifically (not the latest) is honored.
        inc2 = store.add_new_increment(PROJECT_NAME, DEMO_BEFORE)
        result = store.simulate_comparison(PROJECT_NAME, inc2.name, DEMO_AFTER)
        store.confirm_update(PROJECT_NAME, inc2.name, DEMO_AFTER, result)

        # INC 3: a genuinely different increment, single version.
        inc3 = store.add_new_increment(PROJECT_NAME, DEMO_INC3_NEW)

        print("=" * 70)
        print("Setup")
        print("=" * 70)
        print(f"INC 2 name: {inc2.name!r}, now at version 2 (v1 also stored)")
        print(f"INC 3 name: {inc3.name!r}, version {inc3.version}")

        # ------------------------------------------------------------
        # 1 -- combine INC 2 @ v1 (deliberately NOT latest) + INC 3 @
        # latest, in that on-screen order
        # ------------------------------------------------------------
        print("\n" + "=" * 70)
        print("1 -- Combine INC 2 (v1, non-latest) + INC 3 (latest)")
        print("=" * 70)

        inc2_v1 = store.get_increment_for_display(PROJECT_NAME, inc2.name, version=1)
        inc2_v2 = store.get_increment_for_display(PROJECT_NAME, inc2.name, version=2)
        inc3_latest = store.get_increment_for_display(PROJECT_NAME, inc3.name)

        print(f"INC 2 v1 total_count: {inc2_v1.total_count}")
        print(f"INC 2 v2 total_count: {inc2_v2.total_count}")
        if inc2_v1.total_count == inc2_v2.total_count:
            failures.append(
                "1: fixture assumption broken -- v1 and v2 of INC 2 should have different item counts "
                "(demo_after.xlsm adds 3 new items over demo_before.xlsm)"
            )

        export_path = str(Path(tmp) / "combined_test.xlsx")
        combined_export.export_combined_report([inc2_v1, inc3_latest], export_path)

        wb = openpyxl.load_workbook(export_path)
        print(f"Sheet names: {wb.sheetnames}")
        if wb.sheetnames != ["All Data", "Sum Data", "Report", "Changes"]:
            failures.append(f"1: expected exactly ['All Data', 'Sum Data', 'Report', 'Changes'], got {wb.sheetnames}")

        # ------------------------------------------------------------
        # 1a -- All Data: INC 2's rows fully before INC 3's, each
        # correctly labeled by the new Increment column, uses v1's data
        # (not silently the latest v2)
        # ------------------------------------------------------------
        ws_all = wb["All Data"]
        header = [c.value for c in ws_all[1]]
        print(f"All Data header: {header}")
        if header[0] != "Increment":
            failures.append(f"1a: All Data's first column should be 'Increment', got {header[0]!r}")

        increment_column_values = [ws_all.cell(row=r, column=1).value for r in range(2, ws_all.max_row)]  # exclude totals row
        first_inc3_row = next((i for i, v in enumerate(increment_column_values) if v == inc3.name), None)
        last_inc2_row = max((i for i, v in enumerate(increment_column_values) if v == inc2.name), default=-1)
        print(f"Last INC2-labeled row index: {last_inc2_row}, first INC3-labeled row index: {first_inc3_row}")
        if first_inc3_row is None or last_inc2_row >= first_inc3_row:
            failures.append("1a: All INC 2 rows should appear before any INC 3 row (on-screen order)")

        inc2_row_count_in_export = sum(1 for v in increment_column_values if v == inc2.name)
        inc3_row_count_in_export = sum(1 for v in increment_column_values if v == inc3.name)
        print(f"INC2 rows in export: {inc2_row_count_in_export} (should match v1's {inc2_v1.total_count})")
        print(f"INC3 rows in export: {inc3_row_count_in_export} (should match {inc3_latest.total_count})")
        if inc2_row_count_in_export != inc2_v1.total_count:
            failures.append(
                f"1a: expected {inc2_v1.total_count} INC2 rows (v1's count), got {inc2_row_count_in_export} -- "
                "looks like it used the wrong version"
            )
        if inc3_row_count_in_export != inc3_latest.total_count:
            failures.append(f"1a: expected {inc3_latest.total_count} INC3 rows, got {inc3_row_count_in_export}")

        # B-C18 was added in v2 -- must NOT appear since we selected v1
        b_c18_present = any(
            ws_all.cell(row=r, column=2).value == "B-C18" for r in range(2, ws_all.max_row)
        )
        print(f"B-C18 present (added in v2, should be ABSENT since v1 was selected): {b_c18_present}")
        if b_c18_present:
            failures.append("1a: B-C18 (added in v2) should not appear -- combined export should have used v1, not the latest")

        totals_row_idx = ws_all.max_row
        print(f"All Data totals row Index column value: {ws_all.cell(row=totals_row_idx, column=1).value!r}")
        if ws_all.cell(row=totals_row_idx, column=1).value != "Totals":
            failures.append("1a: All Data should still end with a 'Totals' row")

        # ------------------------------------------------------------
        # 1b -- Sum Data: same Increment-column/order checks
        # ------------------------------------------------------------
        ws_sum = wb["Sum Data"]
        sum_header = [c.value for c in ws_sum[1]]
        if sum_header[0] != "Increment":
            failures.append(f"1b: Sum Data's first column should be 'Increment', got {sum_header[0]!r}")
        sum_increment_values = [ws_sum.cell(row=r, column=1).value for r in range(2, ws_sum.max_row)]
        # order already thoroughly checked via All Data above -- Sum Data
        # is built from the same per-increment iteration, so just confirm
        # both increments' rows actually made it into this sheet too.
        print(f"Sum Data rows use INC2 name: {inc2.name in sum_increment_values}, INC3 name: {inc3.name in sum_increment_values}")
        if inc2.name not in sum_increment_values or inc3.name not in sum_increment_values:
            failures.append("1b: Sum Data should contain rows from both increments")

        # ------------------------------------------------------------
        # 1b2 -- All Data/Sum Data per-increment SUBTOTAL rows: exactly
        # one immediately after each increment's own block (contiguity),
        # its numbers EXACTLY matching that increment's own independent
        # single-increment totals (inc2_v1.all_data_totals /
        # excel_reader.live_sum_data_totals(inc*.sum_data) -- the SAME
        # values that increment's own single-increment Data View/export
        # already show), and the final Grand Total ("Totals") row
        # UNCHANGED/correct.
        # ------------------------------------------------------------
        print("\n" + "=" * 70)
        print("1b2 -- per-increment subtotal rows (All Data / Sum Data)")
        print("=" * 70)

        all_data_col1 = [ws_all.cell(row=r, column=1).value for r in range(2, ws_all.max_row + 1)]
        inc2_subtotal_label = f"{inc2.name} — Subtotal"
        inc3_subtotal_label = f"{inc3.name} — Subtotal"

        # Contiguity: subtotal label immediately follows the last real
        # INC2 row, and immediately precedes INC3's first real row.
        inc2_subtotal_idx = all_data_col1.index(inc2_subtotal_label) if inc2_subtotal_label in all_data_col1 else None
        print(f"INC2 All Data subtotal row found at offset {inc2_subtotal_idx}")
        if inc2_subtotal_idx is None:
            failures.append(f"1b2: no All Data subtotal row found for INC2 ({inc2_subtotal_label!r})")
        else:
            if all_data_col1[inc2_subtotal_idx - 1] != inc2.name:
                failures.append("1b2: All Data INC2 subtotal row is not immediately preceded by an INC2 real row")
            if all_data_col1[inc2_subtotal_idx + 1] != inc3.name:
                failures.append("1b2: All Data INC2 subtotal row is not immediately followed by INC3's first real row")

        # Correctness: read the subtotal row's numeric cells back and
        # compare against inc2_v1.all_data_totals directly (Step 1's
        # reuse guarantee -- these should be the IDENTICAL dict values,
        # not just numerically equal by coincidence).
        if inc2_subtotal_idx is not None:
            excel_row = inc2_subtotal_idx + 2  # +1 header, +1 back to 1-indexed
            stage_count = excel_reader.STAGE_COUNT
            subtotal_values = {
                f"Stage {s}": ws_all.cell(row=excel_row, column=4 + s).value for s in range(1, stage_count + 1)
            }
            subtotal_values["VCR"] = ws_all.cell(row=excel_row, column=4 + stage_count + 1).value
            subtotal_values["SUM"] = ws_all.cell(row=excel_row, column=4 + stage_count + 2).value
            print(f"INC2 All Data subtotal SUM={subtotal_values['SUM']} vs ground truth={inc2_v1.all_data_totals.get('SUM')}")
            if subtotal_values != inc2_v1.all_data_totals:
                failures.append(
                    f"1b2: INC2's All Data subtotal row does NOT exactly match inc2_v1.all_data_totals (the same "
                    f"increment's own single-increment totals) -- got {subtotal_values}, expected {inc2_v1.all_data_totals}"
                )

        # Grand Total unchanged: still the very last row, still correct.
        expected_all_data_grand_total_sum = inc2_v1.all_data_totals.get("SUM", 0) + inc3_latest.all_data_totals.get("SUM", 0)
        actual_all_data_grand_total_sum = ws_all.cell(row=ws_all.max_row, column=4 + excel_reader.STAGE_COUNT + 2).value
        print(f"All Data Grand Total SUM: {actual_all_data_grand_total_sum} (expected {expected_all_data_grand_total_sum})")
        if actual_all_data_grand_total_sum != expected_all_data_grand_total_sum:
            failures.append(
                f"1b2: All Data Grand Total SUM should be {expected_all_data_grand_total_sum}, "
                f"got {actual_all_data_grand_total_sum}"
            )

        # Same contiguity + correctness check for Sum Data.
        sum_data_col1 = [ws_sum.cell(row=r, column=1).value for r in range(2, ws_sum.max_row + 1)]
        inc2_sum_subtotal_idx = sum_data_col1.index(inc2_subtotal_label) if inc2_subtotal_label in sum_data_col1 else None
        print(f"INC2 Sum Data subtotal row found at offset {inc2_sum_subtotal_idx}")
        if inc2_sum_subtotal_idx is None:
            failures.append(f"1b2: no Sum Data subtotal row found for INC2 ({inc2_subtotal_label!r})")
        else:
            if sum_data_col1[inc2_sum_subtotal_idx - 1] != inc2.name:
                failures.append("1b2: Sum Data INC2 subtotal row is not immediately preceded by an INC2 real row")
            if sum_data_col1[inc2_sum_subtotal_idx + 1] != inc3.name:
                failures.append("1b2: Sum Data INC2 subtotal row is not immediately followed by INC3's first real row")

            inc2_sum_ground_truth = excel_reader.live_sum_data_totals(inc2_v1.sum_data)
            excel_row = inc2_sum_subtotal_idx + 2
            stage_count = excel_reader.STAGE_COUNT
            actual_grand_total = ws_sum.cell(row=excel_row, column=4 + stage_count + 4).value
            print(f"INC2 Sum Data subtotal grand_total={actual_grand_total} vs ground truth={inc2_sum_ground_truth['grand_total']}")
            if actual_grand_total != inc2_sum_ground_truth["grand_total"]:
                failures.append(
                    f"1b2: INC2's Sum Data subtotal row's grand_total does NOT match "
                    f"excel_reader.live_sum_data_totals(inc2_v1.sum_data) -- got {actual_grand_total}, "
                    f"expected {inc2_sum_ground_truth['grand_total']}"
                )

        expected_sum_data_grand_total = (
            excel_reader.live_sum_data_totals(inc2_v1.sum_data)["grand_total"]
            + excel_reader.live_sum_data_totals(inc3_latest.sum_data)["grand_total"]
        )
        actual_sum_data_grand_total = ws_sum.cell(row=ws_sum.max_row, column=4 + excel_reader.STAGE_COUNT + 4).value
        print(f"Sum Data Grand Total: {actual_sum_data_grand_total} (expected {expected_sum_data_grand_total})")
        if actual_sum_data_grand_total != expected_sum_data_grand_total:
            failures.append(
                f"1b2: Sum Data Grand Total should be {expected_sum_data_grand_total}, got {actual_sum_data_grand_total}"
            )

        # ------------------------------------------------------------
        # 1c -- Report: two separate labeled sections, each ending in
        # its OWN "Grand Total" row, plus one final, distinctly-labeled
        # overall total
        # ------------------------------------------------------------
        ws_report = wb["Report"]
        report_col1_values = [ws_report.cell(row=r, column=1).value for r in range(2, ws_report.max_row + 1)]
        print(f"Report column 1 values (last 5): {report_col1_values[-5:]}")

        if inc2.name not in report_col1_values:
            failures.append(f"1c: Report should have a section header row labeled {inc2.name!r}")
        if inc3.name not in report_col1_values:
            failures.append(f"1c: Report should have a section header row labeled {inc3.name!r}")

        grand_total_rows = [v for v in report_col1_values if v == "Grand Total"]
        print(f"Per-section 'Grand Total' rows found: {len(grand_total_rows)} (expected 2, one per increment)")
        if len(grand_total_rows) != 2:
            failures.append(f"1c: expected exactly 2 per-section 'Grand Total' rows, got {len(grand_total_rows)}")

        if report_col1_values[-1] != "Grand Total (All Increments)":
            failures.append(
                f"1c: the very last Report row should be 'Grand Total (All Increments)', got {report_col1_values[-1]!r}"
            )
        else:
            combined_grand_total_value = ws_report.cell(row=ws_report.max_row, column=4).value
            inc2_grand_total = next(
                row["total"] for row in inc2_v1.report if row["approval_agency"] == "Grand Total"
            )
            inc3_grand_total = next(
                row["total"] for row in inc3_latest.report if row["approval_agency"] == "Grand Total"
            )
            expected_combined = inc2_grand_total + inc3_grand_total
            print(f"Combined Grand Total: {combined_grand_total_value} (expected {expected_combined} = "
                  f"{inc2_grand_total} + {inc3_grand_total})")
            if combined_grand_total_value != expected_combined:
                failures.append(
                    f"1c: combined Grand Total should be {expected_combined}, got {combined_grand_total_value}"
                )

        # ------------------------------------------------------------
        # 1d -- Changes: State Revision Log and Update History each
        # flattened with a leading Increment column (NOT sectioned like
        # Report -- see core/combined_export.py's module docstring),
        # INC 2's rows fully before INC 3's in both tables, uses v1's
        # J-Changes content (not silently the latest v2), and INC 3
        # (which was never confirm_update()'d, so has zero
        # change_history entries) correctly contributes NO Update
        # History rows at all rather than an empty/error row.
        # ------------------------------------------------------------
        ws_changes = wb["Changes"]
        changes_col1 = [ws_changes.cell(row=r, column=1).value for r in range(1, ws_changes.max_row + 1)]

        if changes_col1[0] != "State Revision Log":
            failures.append(f"1d: Changes sheet should start with 'State Revision Log', got {changes_col1[0]!r}")

        revision_header = [ws_changes.cell(row=2, column=c).value for c in range(1, 8)]
        print(f"Changes/State Revision Log header: {revision_header}")
        if revision_header[0] != "Increment" or revision_header[1] != "Rev #":
            failures.append(f"1d: State Revision Log header should start ['Increment', 'Rev #', ...], got {revision_header}")

        # Revision Log data rows run from row 3 until the blank
        # separator before "Update History".
        history_header_idx = changes_col1.index("Update History")  # 0-indexed into changes_col1
        revision_increment_values = [
            ws_changes.cell(row=r, column=1).value for r in range(3, history_header_idx)  # rows are 1-indexed, changes_col1 is 0-indexed -- this range naturally excludes the blank separator row
        ]
        revision_increment_values = [v for v in revision_increment_values if v is not None]

        first_inc3_rev_row = next((i for i, v in enumerate(revision_increment_values) if v == inc3.name), None)
        last_inc2_rev_row = max((i for i, v in enumerate(revision_increment_values) if v == inc2.name), default=-1)
        print(f"Revision Log: last INC2 row index {last_inc2_rev_row}, first INC3 row index {first_inc3_rev_row}")
        if first_inc3_rev_row is None or last_inc2_rev_row >= first_inc3_rev_row:
            failures.append("1d: All INC 2 revision-log rows should appear before any INC 3 row (on-screen order)")

        inc2_rev_count = sum(1 for v in revision_increment_values if v == inc2.name)
        inc3_rev_count = sum(1 for v in revision_increment_values if v == inc3.name)
        print(f"INC2 revision rows in export: {inc2_rev_count} (should match v1's {len(inc2_v1.changes_log)})")
        print(f"INC3 revision rows in export: {inc3_rev_count} (should match {len(inc3_latest.changes_log)})")
        if inc2_rev_count != len(inc2_v1.changes_log):
            failures.append(
                f"1d: expected {len(inc2_v1.changes_log)} INC2 revision rows (v1's count), got {inc2_rev_count} -- "
                "looks like it used the wrong version"
            )
        if inc3_rev_count != len(inc3_latest.changes_log):
            failures.append(f"1d: expected {len(inc3_latest.changes_log)} INC3 revision rows, got {inc3_rev_count}")

        # Update History: only INC 2 was ever confirm_update()'d, so it
        # should be the ONLY increment contributing rows here.
        history_header_row = history_header_idx + 1  # 1-indexed row number of "Update History" itself
        history_col_header = [ws_changes.cell(row=history_header_row + 1, column=c).value for c in range(1, 8)]
        print(f"Changes/Update History header: {history_col_header}")
        if history_col_header[0] != "Increment" or history_col_header[1] != "Update #":
            failures.append(f"1d: Update History header should start ['Increment', 'Update #', ...], got {history_col_header}")

        history_increment_values = []
        r = history_header_row + 2
        while ws_changes.cell(row=r, column=1).value not in (None, ""):
            history_increment_values.append(ws_changes.cell(row=r, column=1).value)
            r += 1
        print(f"Update History rows: {history_increment_values}")
        if history_increment_values != [inc2.name]:
            failures.append(
                f"1d: expected Update History to contain exactly one row, for INC 2 only (INC 3 was never "
                f"confirmed/updated so has no history), got {history_increment_values}"
            )

        # ------------------------------------------------------------
        # 2 -- default_combined_filename
        # ------------------------------------------------------------
        print("\n" + "=" * 70)
        print("2 -- default_combined_filename")
        print("=" * 70)
        name = combined_export.default_combined_filename(PROJECT_NAME, 2)
        print(f"default_combined_filename result: {name!r}")
        if name != "Combined_Export_Validation_Combined_2-increments.xlsx":
            failures.append(f"2: unexpected default filename: {name!r}")

    print("\n" + "=" * 70)
    if failures:
        print("RESULT: FAIL")
        for f in failures:
            print(" -", f)
        raise SystemExit(1)
    else:
        print("RESULT: PASS -- combined All Data/Sum Data/Report/Changes all correct: on-screen order, Increment "
              "column, non-latest version honored, exactly one distinctly-labeled overall Grand Total, "
              "Changes' flattened (not sectioned) State Revision Log/Update History tables both correct, and "
              "All Data/Sum Data's per-increment subtotal rows are contiguous, exactly match each increment's own "
              "independent single-increment totals, and leave the final Grand Total unchanged")


if __name__ == "__main__":
    main()

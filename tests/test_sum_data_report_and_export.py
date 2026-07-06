"""
Validates the Sum Data / Report tabs' underlying data
(ui.mock_data.MockDataStore.get_increment_for_display's sum_data/report
fields) and the 3-sheet Excel export (core.excel_export.export_increment),
using sample_increment.xlsm and a temp-directory ProjectStore (never
touches the real ~/AltamiranoBuildersAppData).

Run directly: `python tests/test_sum_data_report_and_export.py`
"""

import os
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl

from core import excel_export, excel_reader
from core.project_store import ProjectStore
from ui.mock_data import MockDataStore

FIXTURE = os.path.join(os.path.dirname(__file__), "fixtures", "sample_increment.xlsm")
PROJECT_NAME = "Sum Data / Report / Export Validation"


def main():
    failures = []

    with tempfile.TemporaryDirectory() as tmp:
        project_store = ProjectStore(base_dir=Path(tmp) / "data")
        store = MockDataStore(project_store)

        store.add_project(PROJECT_NAME, "/tmp/sdr")
        increment = store.add_new_increment(PROJECT_NAME, FIXTURE)
        display = store.get_increment_for_display(PROJECT_NAME, increment.name)

        wb = excel_reader.open_workbook(FIXTURE)

        # ------------------------------------------------------------
        # 1 -- Sum Data rows: same subset/order build_sum_data() itself
        # selects, and an unset required stage defaults to Open (not
        # blank) when rendered live
        # ------------------------------------------------------------
        print("=" * 70)
        print("1 -- Sum Data rows")
        print("=" * 70)

        expected_sum_data_df = excel_reader.normalize_workbook(wb)["sum_data"]
        expected_indexes = [idx for idx in expected_sum_data_df["Index"] if idx]
        actual_indexes = [row["index"] for row in display.sum_data]
        print(f"Sum Data rows: {len(actual_indexes)} (expected {len(expected_indexes)})")
        if actual_indexes != expected_indexes:
            failures.append(
                "1: Sum Data row set/order should exactly match build_sum_data()'s Index column "
                f"({len(actual_indexes)} rows vs {len(expected_indexes)} expected)"
            )

        # Regression check for the fixed inclusion rule: C-F4 has real
        # required stages (10, 11, 12, 17, 18, 20 -- all "1"/Open) but a
        # blank Approval Agency, and used to be wrongly excluded from Sum
        # Data on that basis alone (see core/excel_reader.py's SUM DATA
        # section, "*** BUG FIXED ***"). It must now appear.
        if "C-F4" not in actual_indexes:
            failures.append("1: C-F4 has real required stages and should now appear in Sum Data despite its blank Approval Agency")
        else:
            cf4_row = next(row for row in display.sum_data if row["index"] == "C-F4")
            print(f"C-F4 required_stages: {cf4_row['required_stages']}")
            if not cf4_row["required_stages"]:
                failures.append("1: C-F4 appears in Sum Data but has no required_stages -- fixture assumption broken")

        sample_row = next(row for row in display.sum_data if row["required_stages"])
        sample_index = sample_row["index"]
        sample_stage = sample_row["required_stages"][0]
        print(f"Sample row {sample_index!r}, stage {sample_stage}")

        # Simulate "required but never explicitly marked" by clearing
        # just this one (index, stage)'s status.json entry directly (the
        # raw-value-trust seeding on add_new_increment means every
        # required stage already has an explicit entry, so this is the
        # only way to exercise the default-to-Open fallback).
        status_map = project_store.load_status(PROJECT_NAME, increment.name)
        del status_map[sample_index][sample_stage]
        project_store.save_status(PROJECT_NAME, increment.name, status_map)

        display2 = store.get_increment_for_display(PROJECT_NAME, increment.name)
        sample_row2 = next(row for row in display2.sum_data if row["index"] == sample_index)
        live_status = sample_row2["stage_status"].get(sample_stage, "Open")
        if live_status != "Open":
            failures.append(f"1: an unset required stage should default to Open in Sum Data, got {live_status!r}")

        # ------------------------------------------------------------
        # 2 -- Sum Data rows share the SAME stage_status dict object as
        # the matching All Data row -- an edit via set_stage_status
        # should be visible on both without re-loading
        # ------------------------------------------------------------
        print("\n" + "=" * 70)
        print("2 -- Sum Data shares live status with All Data")
        print("=" * 70)

        all_data_row = next(row for row in display2.all_data if row["index"] == sample_index)
        sum_data_row = next(row for row in display2.sum_data if row["index"] == sample_index)
        if all_data_row["stage_status"] is not sum_data_row["stage_status"]:
            failures.append(
                "2: All Data and Sum Data rows for the same index should share the same "
                "stage_status dict object"
            )

        store.set_stage_status(PROJECT_NAME, increment.name, sample_index, sample_stage, "Done")
        all_data_row["stage_status"][sample_stage] = "Done"  # mirror the in-memory update the click handler makes
        if sum_data_row["stage_status"].get(sample_stage) != "Done":
            failures.append("2: editing the shared dict via the All Data row should be visible from the Sum Data row")

        # ------------------------------------------------------------
        # 3 -- Report: matches build_report(), Grand Total is the
        # arithmetically correct sum, not the ground-truth file's
        # documented doubled-total bug (590 -- see
        # core/excel_reader.py's REPORT module docstring section)
        # ------------------------------------------------------------
        print("\n" + "=" * 70)
        print("3 -- Report rows and Grand Total")
        print("=" * 70)

        expected_report_df = excel_reader.normalize_workbook(wb)["report"]
        if len(display.report) != len(expected_report_df):
            failures.append(
                f"3: Report row count should match build_report() ({len(display.report)} vs "
                f"{len(expected_report_df)})"
            )

        grand_total_rows = [row for row in display.report if row["approval_agency"] == "Grand Total"]
        if len(grand_total_rows) != 1:
            failures.append(f"3: Report should have exactly one Grand Total row, got {len(grand_total_rows)}")
        else:
            grand_total = grand_total_rows[0]["total"]
            individual_sum = sum(row["total"] for row in display.report if row["approval_agency"] != "Grand Total")
            print(f"Grand Total: {grand_total}, sum of individual rows: {individual_sum}")
            if grand_total != individual_sum:
                failures.append(
                    f"3: Grand Total ({grand_total}) should equal the sum of individual row "
                    f"totals ({individual_sum})"
                )
            if grand_total == 590:
                failures.append("3: Grand Total should not replicate the ground-truth file's documented duplicate-row bug (590)")

        # continuation rows (same Approval Agency as the row above --
        # see core.excel_reader.build_report()) must show None, never
        # the literal string "nan" -- pandas can coerce that column's
        # None placeholders to float NaN, which is truthy in Python and
        # slips past a plain `or ""` check (see ui.mock_data._none_if_nan)
        nan_like = [row for row in display.report if str(row.get("approval_agency")) == "nan"]
        if nan_like:
            failures.append(
                f"3: {len(nan_like)} Report row(s) show literal 'nan' for approval_agency instead "
                f"of None -- e.g. {nan_like[0]}"
            )
        continuation_rows = [row for row in display.report if row["approval_agency"] is None]
        print(f"Continuation rows (approval_agency is None): {len(continuation_rows)}")
        if not continuation_rows:
            failures.append("3: expected at least one continuation row (same agency as the row above) with approval_agency is None")

        # ------------------------------------------------------------
        # 4 -- Export: 3 sheets, right names/order, headers bolded,
        # freeze panes set, data matches what's in the app
        # ------------------------------------------------------------
        print("\n" + "=" * 70)
        print("4 -- Export to Excel")
        print("=" * 70)

        export_path = str(Path(tmp) / "export_test.xlsx")
        excel_export.export_increment(display2, export_path)

        wb_out = openpyxl.load_workbook(export_path)
        print(f"Sheet names: {wb_out.sheetnames}")
        if wb_out.sheetnames != ["All Data", "Sum Data", "Report"]:
            failures.append(f"4: sheet names/order should be exactly ['All Data', 'Sum Data', 'Report'], got {wb_out.sheetnames}")

        ws_all = wb_out["All Data"]
        if not ws_all["A1"].font.bold:
            failures.append("4: All Data header row should be bold")
        if ws_all.freeze_panes != "D2":
            failures.append(f"4: All Data should freeze header row + first 3 columns (D2), got {ws_all.freeze_panes}")
        if ws_all.max_row - 1 != len(display2.all_data):
            failures.append(f"4: All Data sheet should have one row per item ({ws_all.max_row - 1} vs {len(display2.all_data)})")

        ws_sum = wb_out["Sum Data"]
        if not ws_sum["A1"].font.bold:
            failures.append("4: Sum Data header row should be bold")
        if ws_sum.freeze_panes != "D2":
            failures.append(f"4: Sum Data should freeze header row + first 3 columns (D2), got {ws_sum.freeze_panes}")
        if ws_sum.max_row - 1 != len(display2.sum_data):
            failures.append(f"4: Sum Data sheet should have one row per item ({ws_sum.max_row - 1} vs {len(display2.sum_data)})")
        if "% Complete" not in [c.value for c in ws_sum[1]]:
            failures.append("4: Sum Data sheet should have a % Complete column")

        ws_report = wb_out["Report"]
        if not ws_report["A1"].font.bold:
            failures.append("4: Report header row should be bold")
        if ws_report.max_row - 1 != len(display2.report):
            failures.append(f"4: Report sheet should have one row per group + Grand Total ({ws_report.max_row - 1} vs {len(display2.report)})")

        sample_row_in_export = next(
            r for r in range(2, ws_all.max_row + 1) if ws_all.cell(row=r, column=1).value == sample_index
        )
        sample_cell = ws_all.cell(row=sample_row_in_export, column=3 + sample_stage)
        print(f"Exported All Data cell for {sample_index} stage {sample_stage}: value={sample_cell.value!r}")
        if sample_cell.value != "X":
            failures.append(
                f"4: exported All Data cell for {sample_index} stage {sample_stage} should be "
                f"'X' (Done), got {sample_cell.value!r}"
            )

        # ------------------------------------------------------------
        # 4b -- Description cells wrap: descriptions carry embedded \n
        # (component / code citation / test name, 3 logical lines) that
        # render as jammed-together text in Excel unless wrap_text is on
        # (see core/excel_export.py's _apply_description_wrap). Checked
        # on every data row of All Data and Sum Data, not just a sample.
        # ------------------------------------------------------------
        print("\n" + "=" * 70)
        print("4b -- Description cells have wrap_text enabled")
        print("=" * 70)

        for sheet_name, ws in (("All Data", ws_all), ("Sum Data", ws_sum)):
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=excel_export.DESCRIPTION_COL)
                if not cell.alignment.wrap_text:
                    failures.append(
                        f"4b: {sheet_name}!{cell.coordinate} (Description) should have "
                        f"wrap_text=True, got {cell.alignment.wrap_text!r}"
                    )
                    break
            row_height = ws.row_dimensions[2].height
            if not row_height or row_height < 30:
                failures.append(
                    f"4b: {sheet_name} data rows should have a row height tall enough for a "
                    f"3-line description, got {row_height!r} for row 2"
                )

        description_width = ws_all.column_dimensions[
            openpyxl.utils.get_column_letter(excel_export.DESCRIPTION_COL)
        ].width
        print(f"Description column width: {description_width}")
        if not (45 <= description_width <= 55):
            failures.append(f"4b: Description column width should be 45-55, got {description_width!r}")

        # ------------------------------------------------------------
        # 5 -- default_filename sanitizes filesystem-unsafe characters
        # ------------------------------------------------------------
        print("\n" + "=" * 70)
        print("5 -- default_filename sanitization")
        print("=" * 70)
        name = excel_export.default_filename('A/B: Project?', 'INC "2" <test>', 3)
        print(f"default_filename result: {name!r}")
        if any(c in name for c in '\\/:*?"<>|'):
            failures.append(f"5: default_filename should strip filesystem-unsafe characters, got {name!r}")
        if not name.endswith("_v3.xlsx"):
            failures.append(f"5: default_filename should end with '_v{{version}}.xlsx', got {name!r}")

        # ------------------------------------------------------------
        # 6 -- demo_before.xlsm -> demo_after.xlsm: Sum Data grows to
        # include the 3 new items (B-C18/19/20), matching All Data's
        # item-count growth -- rather than staying fixed at whatever
        # row count the first upload produced. This is the concrete bug
        # report that led to the SUM DATA inclusion-rule fix above.
        # ------------------------------------------------------------
        print("\n" + "=" * 70)
        print("6 -- Sum Data grows on demo_before.xlsm -> demo_after.xlsm")
        print("=" * 70)

        demo_before = os.path.join(os.path.dirname(__file__), "fixtures", "demo_before.xlsm")
        demo_after = os.path.join(os.path.dirname(__file__), "fixtures", "demo_after.xlsm")
        growth_project = "Sum Data Growth Validation"

        store.add_project(growth_project, "/tmp/growth")
        demo_increment = store.add_new_increment(growth_project, demo_before)
        display_before = store.get_increment_for_display(growth_project, demo_increment.name)

        result = store.simulate_comparison(growth_project, demo_increment.name, demo_after)
        store.confirm_update(growth_project, demo_increment.name, demo_after, result)
        display_after = store.get_increment_for_display(growth_project, demo_increment.name)

        all_data_growth = display_after.total_count - display_before.total_count
        sum_data_growth = len(display_after.sum_data) - len(display_before.sum_data)
        print(f"All Data grew by {all_data_growth} items, Sum Data grew by {sum_data_growth} items")
        if all_data_growth != 3:
            failures.append(
                f"6: fixture assumption broken -- expected demo_after.xlsm to add exactly 3 items "
                f"to All Data, got {all_data_growth}"
            )
        if sum_data_growth != all_data_growth:
            failures.append(
                f"6: Sum Data should grow by the same 3 items All Data does, not stay fixed -- "
                f"All Data grew by {all_data_growth}, Sum Data grew by {sum_data_growth}"
            )

        after_indexes = {row["index"] for row in display_after.sum_data}
        missing_from_sum_data = {"B-C18", "B-C19", "B-C20"} - after_indexes
        print(f"B-C18/19/20 present in Sum Data after update: {not missing_from_sum_data}")
        if missing_from_sum_data:
            failures.append(
                f"6: {sorted(missing_from_sum_data)} should appear in Sum Data after uploading "
                f"demo_after.xlsm, but don't"
            )

    print("\n" + "=" * 70)
    print("SUMMARY")
    print("=" * 70)
    if failures:
        print("RESULT: FAIL")
        for f in failures:
            print(" -", f)
        raise SystemExit(1)
    else:
        print("RESULT: PASS")


if __name__ == "__main__":
    main()

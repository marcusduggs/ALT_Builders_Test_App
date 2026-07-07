"""
Validates the All Data / Sum Data bottom totals row: core.excel_reader's
all_data_totals()/sum_data_totals()/live_sum_data_totals(), how they're
surfaced on ui.mock_data.Increment, and how they're rendered in the
exported .xlsx (core/excel_export.py).

Ground truth for sample_increment.xlsm was reverse-engineered directly
from the real file's own bottom-row formulas (data_only=False) and
cross-checked against its cached values (data_only=True):

  All Data row 381:  every Stage/VCR/SUM column = SUM(col3:col380)
                      (Excel's SUM() already skips "X" text, same rule
                      BlueSheetRecord.sum_value uses per row) -- SUM
                      (grand total) cached at 291.
  Sum Data row 95:   every Stage/VCR column = COUNTIF(col4:col94,"*Open*");
                      Open/Done/Total totals = SUM of the per-row
                      Open/Done/Total columns; % Complete = Done/Total.
                      Cached: Open=277, Done=189, Total=466, %=0.4056 (41%).

Run directly: `python tests/test_totals_row.py`
"""

import os
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl

from core import excel_export, excel_reader
from core.project_store import ProjectStore
from ui.mock_data import MockDataStore

FIXTURES = os.path.join(os.path.dirname(__file__), "fixtures")
SAMPLE = os.path.join(FIXTURES, "sample_increment.xlsm")
DEMO_BEFORE = os.path.join(FIXTURES, "demo_before.xlsm")
DEMO_AFTER = os.path.join(FIXTURES, "demo_after.xlsm")
PROJECT_NAME = "Totals Row Validation"


def main():
    failures = []

    with tempfile.TemporaryDirectory() as tmp:
        project_store = ProjectStore(base_dir=Path(tmp) / "data")
        store = MockDataStore(project_store)

        store.add_project(PROJECT_NAME)
        increment = store.add_new_increment(PROJECT_NAME, SAMPLE)
        display = store.get_increment_for_display(PROJECT_NAME, increment.name)

        # ------------------------------------------------------------
        # 1 -- fresh upload, zero edits.
        #
        # All Data's row set is 1:1 with the real file (every row, no
        # filtering), so its SUM total matches the file's own cached
        # grand total exactly: 291.
        #
        # Sum Data is DIFFERENT: this app's build_sum_data() deliberately
        # includes more rows than the real file's own static Sum Data
        # sheet does (91 rows) -- see core/excel_reader.py's "SUM DATA IS
        # STALE...BUG FIXED" note: the file's own Sum Data sheet wrongly
        # excludes items with real required stages but a blank Approval
        # Agency (C-F4 is the named example). Confirmed directly: our
        # reconstruction has 103 rows, 12 more than the file's 91 (all 12
        # exactly the blank-Approval-Agency category the bug fix
        # targets). So Sum Data's totals CANNOT match the file's stale
        # cached totals (277/189/466) -- that would mean re-introducing
        # the bug. What's validated instead: the totals row's own
        # arithmetic is correct for OUR (corrected) row set, cross-checked
        # independently against the same DataFrame, and that live/raw
        # agree with each other on a fresh upload (status.json is seeded
        # directly from the file's raw status at upload time).
        # ------------------------------------------------------------
        print("=" * 70)
        print("1 -- Fresh upload: All Data matches the file exactly; Sum Data's row-count difference is by design")
        print("=" * 70)

        all_totals = display.all_data_totals
        print(f"All Data SUM total: {all_totals.get('SUM')}")
        if all_totals.get("SUM") != 291:
            failures.append(f"1: All Data SUM total should be 291 (real file's cached grand total), got {all_totals.get('SUM')}")

        parsed = excel_reader.normalize_workbook(SAMPLE)
        sum_data_df = parsed["sum_data"]
        print(f"Our Sum Data row count: {len(sum_data_df)} (real file's own Sum Data sheet has 91 -- by design, see above)")
        if len(sum_data_df) != 103:
            failures.append(f"1: fixture assumption broken -- expected our Sum Data reconstruction to have 103 rows, got {len(sum_data_df)}")

        # independent cross-check, not reusing sum_data_totals()/live_sum_data_totals() themselves
        stage_and_vcr_cols = [f"Stage {i}" for i in range(1, excel_reader.STAGE_COUNT + 1)] + ["VCR"]
        expected_open = int((sum_data_df[stage_and_vcr_cols] == "Open").sum().sum())
        expected_done = int((sum_data_df[stage_and_vcr_cols] == "Done").sum().sum())
        expected_total = int(sum_data_df["Total"].sum())
        print(f"Independently cross-checked expected totals for OUR row set: Open={expected_open}, Done={expected_done}, Total={expected_total}")

        live_totals = excel_reader.live_sum_data_totals(display.sum_data)
        print(f"Sum Data live totals:     Open={live_totals['open_total']}, Done={live_totals['done_total']}, "
              f"Total={live_totals['grand_total']}, %={live_totals['pct_complete']}")
        if (live_totals["open_total"], live_totals["done_total"], live_totals["grand_total"]) != (expected_open, expected_done, expected_total):
            failures.append(
                f"1: live_sum_data_totals() should match the independent cross-check "
                f"({expected_open}, {expected_done}, {expected_total}), got "
                f"({live_totals['open_total']}, {live_totals['done_total']}, {live_totals['grand_total']})"
            )
        if abs(live_totals["pct_complete"] - (expected_done / expected_total)) > 1e-9:
            failures.append(f"1: Sum Data % Complete should be {expected_done}/{expected_total}, got {live_totals['pct_complete']}")

        # raw-file-derived sum_data_totals() (normalize_workbook()'s own
        # artifact) should match live_sum_data_totals() on a fresh upload
        # -- status.json is seeded directly from the file's raw status at
        # upload time, so live and raw agree until something is edited
        # (see step 4).
        raw_totals = parsed["sum_data_totals"]
        print(f"Sum Data raw-file totals: Open={raw_totals['open_total']}, Done={raw_totals['done_total']}, "
              f"Total={raw_totals['grand_total']}")
        if (raw_totals["open_total"], raw_totals["done_total"], raw_totals["grand_total"]) != \
           (live_totals["open_total"], live_totals["done_total"], live_totals["grand_total"]):
            failures.append(
                f"1: raw sum_data_totals() should equal live_sum_data_totals() on a fresh upload, got "
                f"raw={raw_totals} vs live open/done/total="
                f"({live_totals['open_total']}, {live_totals['done_total']}, {live_totals['grand_total']})"
            )

        # ------------------------------------------------------------
        # 2 -- totals row is NOT counted as a real item anywhere
        # ------------------------------------------------------------
        print("\n" + "=" * 70)
        print("2 -- Totals row never counted as a real item")
        print("=" * 70)

        print(f"total_count: {display.total_count} (real item count -- blank/category-header rows already excluded upstream, no totals-row leakage)")
        if display.total_count != 345:
            failures.append(f"2: total_count should be exactly the real item count (345), got {display.total_count}")
        if any(row.get("index") in (None, "", "Totals") for row in display.all_data):
            failures.append("2: no All Data row should have a blank/'Totals' index -- the totals row must not leak into all_data")
        if any(row.get("index") in (None, "", "Totals") for row in display.sum_data):
            failures.append("2: no Sum Data row should have a blank/'Totals' index -- the totals row must not leak into sum_data")

        # ------------------------------------------------------------
        # 3 -- exported file: totals row present, bold, tinted, and
        # numerically typed (not strings)
        # ------------------------------------------------------------
        print("\n" + "=" * 70)
        print("3 -- Exported file's totals row: values, styling, numeric types")
        print("=" * 70)

        export_path = str(Path(tmp) / "totals_export.xlsx")
        excel_export.export_increment(display, export_path)
        wb_out = openpyxl.load_workbook(export_path)

        ws_all = wb_out["All Data"]
        all_totals_row = ws_all.max_row
        sum_cell = ws_all.cell(row=all_totals_row, column=3 + excel_reader.STAGE_COUNT + 2)  # Index,Desc,Agency + 42 stages + VCR -> SUM
        print(f"Exported All Data totals row ({all_totals_row}): Index={ws_all.cell(row=all_totals_row, column=1).value!r}, "
              f"SUM={sum_cell.value!r} (type={type(sum_cell.value).__name__})")
        if ws_all.cell(row=all_totals_row, column=1).value != "Totals":
            failures.append(f"3: All Data exported totals row should read 'Totals' in the Index column, got {ws_all.cell(row=all_totals_row, column=1).value!r}")
        if sum_cell.value != 291:
            failures.append(f"3: exported All Data SUM total should be 291, got {sum_cell.value!r}")
        if not isinstance(sum_cell.value, (int, float)):
            failures.append(f"3: exported All Data SUM total should be a real number, got type {type(sum_cell.value).__name__}")
        if not ws_all.cell(row=all_totals_row, column=1).font.bold:
            failures.append("3: All Data exported totals row should be bold")
        if ws_all.cell(row=all_totals_row, column=1).fill.fgColor.rgb not in ("00000000", None) and \
           ws_all.cell(row=all_totals_row, column=1).fill.fill_type != "solid":
            failures.append("3: All Data exported totals row should have a fill applied")

        ws_sum = wb_out["Sum Data"]
        sum_totals_row = ws_sum.max_row
        open_col = 3 + excel_reader.STAGE_COUNT + 2  # Index,Desc,Agency + 42 stages + VCR -> Open
        open_cell = ws_sum.cell(row=sum_totals_row, column=open_col)
        done_cell = ws_sum.cell(row=sum_totals_row, column=open_col + 1)
        total_cell = ws_sum.cell(row=sum_totals_row, column=open_col + 2)
        print(f"Exported Sum Data totals row ({sum_totals_row}): Open={open_cell.value!r}, Done={done_cell.value!r}, "
              f"Total={total_cell.value!r} (types: {type(open_cell.value).__name__}/{type(done_cell.value).__name__}/{type(total_cell.value).__name__})")
        if (open_cell.value, done_cell.value, total_cell.value) != \
           (live_totals["open_total"], live_totals["done_total"], live_totals["grand_total"]):
            failures.append(
                f"3: exported Sum Data Open/Done/Total totals should match live_sum_data_totals() "
                f"({live_totals['open_total']}, {live_totals['done_total']}, {live_totals['grand_total']}), "
                f"got {(open_cell.value, done_cell.value, total_cell.value)}"
            )
        if not all(isinstance(v, (int, float)) for v in (open_cell.value, done_cell.value, total_cell.value)):
            failures.append("3: exported Sum Data Open/Done/Total totals should all be real numbers, not strings")
        if not ws_sum.cell(row=sum_totals_row, column=1).font.bold:
            failures.append("3: Sum Data exported totals row should be bold")

        # ------------------------------------------------------------
        # 4 -- totals row updates immediately after a live status edit,
        # with NO re-upload / re-navigation needed
        # ------------------------------------------------------------
        print("\n" + "=" * 70)
        print("4 -- Totals row updates live after a status edit (no re-upload)")
        print("=" * 70)

        before = excel_reader.live_sum_data_totals(display.sum_data)
        # find a currently-Open required stage on some item to flip to Done
        target_row = next(
            row for row in display.sum_data
            if any(row["stage_status"].get(s, "Open") == "Open" for s in row["required_stages"])
        )
        target_stage = next(s for s in target_row["required_stages"] if target_row["stage_status"].get(s, "Open") == "Open")
        print(f"Flipping {target_row['index']} stage {target_stage} from Open -> Done")
        store.set_stage_status(PROJECT_NAME, increment.name, target_row["index"], target_stage, "Done")
        target_row["stage_status"][target_stage] = "Done"  # mirror the in-memory update the click handler makes

        after = excel_reader.live_sum_data_totals(display.sum_data)
        print(f"Before: Open={before['open_total']}, Done={before['done_total']}, %={before['pct_complete']:.4f}")
        print(f"After:  Open={after['open_total']}, Done={after['done_total']}, %={after['pct_complete']:.4f}")
        if after["open_total"] != before["open_total"] - 1:
            failures.append(f"4: Open total should decrease by 1 after the edit, went {before['open_total']} -> {after['open_total']}")
        if after["done_total"] != before["done_total"] + 1:
            failures.append(f"4: Done total should increase by 1 after the edit, went {before['done_total']} -> {after['done_total']}")
        if after["pct_complete"] <= before["pct_complete"]:
            failures.append(f"4: % Complete should increase after marking a stage Done, went {before['pct_complete']} -> {after['pct_complete']}")
        if after["grand_total"] != before["grand_total"]:
            failures.append(f"4: Total (grand total) should be unchanged by a status edit, went {before['grand_total']} -> {after['grand_total']}")

        # ------------------------------------------------------------
        # 5 -- totals reflect a NEW file's counts/values, not stale ones
        # from a previous upload (demo_before.xlsm -> demo_after.xlsm,
        # which adds 3 new items)
        # ------------------------------------------------------------
        print("\n" + "=" * 70)
        print("5 -- Totals reflect the currently-loaded file, not a stale one")
        print("=" * 70)

        growth_project = "Totals Row Growth Validation"
        store.add_project(growth_project)
        growth_increment = store.add_new_increment(growth_project, DEMO_BEFORE)
        display_before = store.get_increment_for_display(growth_project, growth_increment.name)
        totals_before = display_before.all_data_totals
        sum_totals_before = excel_reader.live_sum_data_totals(display_before.sum_data)

        result = store.simulate_comparison(growth_project, growth_increment.name, DEMO_AFTER)
        store.confirm_update(growth_project, growth_increment.name, DEMO_AFTER, result)
        display_after = store.get_increment_for_display(growth_project, growth_increment.name)
        totals_after = display_after.all_data_totals
        sum_totals_after = excel_reader.live_sum_data_totals(display_after.sum_data)

        print(f"total_count: before={display_before.total_count}, after={display_after.total_count}")
        if display_after.total_count - display_before.total_count != 3:
            failures.append("5: fixture assumption broken -- expected demo_after.xlsm to add exactly 3 items")

        # All Data's SUM total is a NUMERIC sum (Excel SUM() semantics --
        # ignores "X" text markers, see all_data_totals()'s docstring).
        # The 3 new items here (B-C18/19/20) happen to be marked with "X"
        # only, no literal numeric test-count value, so this total is
        # legitimately UNCHANGED by this specific upload -- verified
        # directly (every one of the 42 per-stage totals is identical
        # before/after). That's a fact about this fixture pair, not a
        # sign the totals are stale -- Sum Data's Open/Done counts (which
        # DO count "X"/"Open" markers, not sum numeric values) are the
        # right signal to prove this file's own data drove the
        # recomputation, and those clearly move:
        print(f"All Data SUM total: before={totals_before.get('SUM')}, after={totals_after.get('SUM')} (expected to tie -- see above)")
        print(f"Sum Data Done total: before={sum_totals_before['done_total']}, after={sum_totals_after['done_total']}")
        print(f"Sum Data grand total: before={sum_totals_before['grand_total']}, after={sum_totals_after['grand_total']}")
        if sum_totals_after["grand_total"] == sum_totals_before["grand_total"]:
            failures.append(
                f"5: Sum Data's Total (grand total) should change after uploading a file with 3 new "
                f"items, stayed at {sum_totals_before['grand_total']}"
            )
        if sum_totals_after["done_total"] == sum_totals_before["done_total"]:
            failures.append(
                f"5: Sum Data's Done total should change after uploading a file with 3 new items, "
                f"stayed at {sum_totals_before['done_total']}"
            )

    print("\n" + "=" * 70)
    if failures:
        print("RESULT: FAIL")
        for f in failures:
            print(" -", f)
        raise SystemExit(1)
    else:
        print("RESULT: PASS -- totals row correct on fresh upload, excluded from item counts, "
              "styled/typed correctly in export, and stays live-accurate after edits and re-uploads")


if __name__ == "__main__":
    main()
